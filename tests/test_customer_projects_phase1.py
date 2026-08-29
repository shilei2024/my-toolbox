"""Phase 1 customer-project domain, permission, API and page regression tests."""
from __future__ import annotations

import os
import hashlib
import json
import time
import unittest
from datetime import datetime, timedelta, timezone
from decimal import Decimal
from io import BytesIO

from openpyxl import load_workbook

os.environ["FLASK_ENV"] = "production"
os.environ["DATABASE_URL"] = "sqlite://"
os.environ["SECRET_KEY"] = "phase-1-test-secret-key-that-is-long-enough"
os.environ["ADMIN_EMAIL"] = "bootstrap@test.com"
os.environ["ADMIN_PASSWORD"] = "SafeBootstrapPassword123!"

from app import app  # noqa: E402
from customer_projects.models import (  # noqa: E402
    Customer,
    CustomerProject,
    MaterialCompetitor,
    ProjectExportPolicy,
    ProjectSavedView,
    ProjectActivity,
    ProjectComment,
    ProjectCommentMention,
    ProjectMaterial,
    ProjectMember,
    ProjectStageEvent,
)
from customer_projects.services.projects import (  # noqa: E402
    bootstrap_organization,
    build_market_scope,
    create_customer,
    create_project,
    local_day_bounds,
)
from extensions import db  # noqa: E402
from models import User  # noqa: E402
from shared.models import AuditEvent, Organization, OrganizationMembership  # noqa: E402


class CustomerProjectsPhase1Test(unittest.TestCase):
    def setUp(self) -> None:
        app.config.update(
            TESTING=True,
            WTF_CSRF_ENABLED=False,
            CUSTOMER_PROJECTS_ENABLED=True,
            CUSTOMER_PROJECTS_PILOT_EMAILS="",
        )
        with app.app_context():
            db.drop_all()
            db.create_all()
            self.admin_id = self._add_user("manager@test.com", admin=True)
            self.sales_id = self._add_user("sales@test.com")
            self.other_id = self._add_user("other@test.com")
            self.fae_id = self._add_user("fae@test.com")
            org = bootstrap_organization("测试组织", self.admin_id)
            self.org_id = org.id
            sales_membership = OrganizationMembership(
                organization_id=org.id, user_id=self.sales_id
            )
            sales_membership.set_roles(["sales"])
            db.session.add(sales_membership)
            other_membership = OrganizationMembership(
                organization_id=org.id, user_id=self.other_id
            )
            other_membership.set_roles(["sales"])
            db.session.add(other_membership)
            fae_membership = OrganizationMembership(
                organization_id=org.id, user_id=self.fae_id
            )
            fae_membership.set_roles(["fae"])
            db.session.add(fae_membership)
            db.session.commit()
        self.client = app.test_client()

    def _add_user(self, email: str, admin: bool = False) -> int:
        user = User(email=email, is_admin=admin, is_active_user=True)
        user.set_password("Passw0rd1")
        db.session.add(user)
        db.session.flush()
        return user.id

    def _login(self, email: str = "sales@test.com") -> None:
        response = self.client.post(
            "/login", data={"email": email, "password": "Passw0rd1"}
        )
        self.assertIn(response.status_code, (301, 302))

    def _seed_project(self) -> str:
        with app.app_context():
            membership = db.session.scalar(
                db.select(OrganizationMembership).where(
                    OrganizationMembership.user_id == self.sales_id
                )
            )
            customer = create_customer({"name": "示例电子"}, membership)
            db.session.flush()
            project = create_project(
                {
                    "customer_id": customer.id,
                    "name": "车载控制器",
                    "product_name": "车载电源控制器",
                    "annual_usage": "120000",
                    "stage_code": "evaluation",
                    "primary_sales_user_id": self.sales_id,
                    "next_action": "确认样品数量",
                    "next_follow_up_at": (
                        datetime.now(timezone.utc) + timedelta(days=2)
                    ).isoformat(),
                    "assessment_grade": "B",
                    "probability_band": 30,
                },
                membership,
                "seed-project",
            )
            db.session.commit()
            return project.id

    def test_feature_flag_hides_module_and_navigation(self) -> None:
        self._login()
        app.config["CUSTOMER_PROJECTS_ENABLED"] = False
        self.assertEqual(self.client.get("/customer-projects/").status_code, 404)
        home = self.client.get("/").get_data(as_text=True)
        self.assertNotIn("客户项目</a>", home)

    def test_project_create_is_idempotent_and_scoped(self) -> None:
        project_id = self._seed_project()
        self._login()
        listing = self.client.get("/api/v1/customer-projects/projects")
        self.assertEqual(listing.status_code, 200)
        self.assertEqual(listing.get_json()["data"][0]["id"], project_id)
        detail = self.client.get(f"/api/v1/customer-projects/projects/{project_id}")
        self.assertEqual(detail.headers["ETag"], '"1"')
        self.assertEqual(detail.get_json()["data"]["product_name"], "车载电源控制器")
        self.assertEqual(detail.get_json()["data"]["annual_usage"], "120000")
        with app.app_context():
            self.assertEqual(db.session.query(CustomerProject).count(), 1)
            self.assertEqual(db.session.query(ProjectStageEvent).count(), 1)

    def test_optimistic_lock_returns_409_without_overwrite(self) -> None:
        project_id = self._seed_project()
        self._login()
        first = self.client.patch(
            f"/api/v1/customer-projects/projects/{project_id}",
            json={"next_action": "准备 20 片样品"},
            headers={"If-Match": '"1"'},
        )
        self.assertEqual(first.status_code, 200)
        self.assertEqual(first.headers["ETag"], '"2"')
        stale = self.client.patch(
            f"/api/v1/customer-projects/projects/{project_id}",
            json={"next_action": "覆盖他人内容"},
            headers={"If-Match": '"1"'},
        )
        self.assertEqual(stale.status_code, 409)
        self.assertEqual(stale.get_json()["error"]["code"], "PROJECT_VERSION_CONFLICT")
        with app.app_context():
            self.assertEqual(db.session.get(CustomerProject, project_id).next_action, "准备 20 片样品")

    def test_activity_updates_snapshot_and_is_idempotent(self) -> None:
        project_id = self._seed_project()
        self._login()
        payload = {
            "activity_type": "meeting",
            "summary": "确认送样计划",
            "next_action": "安排寄送",
            "next_follow_up_at": (datetime.now(timezone.utc) + timedelta(days=4)).isoformat(),
            "project_version": 1,
            "is_meaningful": True,
        }
        first = self.client.post(
            f"/api/v1/customer-projects/projects/{project_id}/activities",
            json=payload,
            headers={"Idempotency-Key": "activity-one"},
        )
        self.assertEqual(first.status_code, 201)
        second = self.client.post(
            f"/api/v1/customer-projects/projects/{project_id}/activities",
            json=payload,
            headers={"Idempotency-Key": "activity-one"},
        )
        self.assertEqual(second.status_code, 201)
        with app.app_context():
            self.assertEqual(db.session.query(ProjectActivity).count(), 1)
            project = db.session.get(CustomerProject, project_id)
            self.assertEqual(project.version, 2)
            self.assertEqual(project.next_action, "安排寄送")

    def test_timeline_comment_mentions_member_without_changing_project_version(self) -> None:
        project_id = self._seed_project()
        self._login()
        payload = {
            "body": "请协助确认客户测试结论。",
            "mention_user_ids": [self.other_id, self.fae_id],
        }
        first = self.client.post(
            f"/api/v1/customer-projects/projects/{project_id}/comments",
            json=payload,
            headers={"Idempotency-Key": "comment-one"},
        )
        self.assertEqual(first.status_code, 201)
        self.assertEqual(first.get_json()["data"]["mention_user_ids"], [self.other_id, self.fae_id])
        retry = self.client.post(
            f"/api/v1/customer-projects/projects/{project_id}/comments",
            json=payload,
            headers={"Idempotency-Key": "comment-one"},
        )
        self.assertEqual(retry.status_code, 201)
        with app.app_context():
            self.assertEqual(db.session.query(ProjectComment).count(), 1)
            self.assertEqual(db.session.query(ProjectCommentMention).count(), 2)
            self.assertEqual(db.session.get(CustomerProject, project_id).version, 1)
        page = self.client.get(f"/customer-projects/projects/{project_id}")
        html = page.get_data(as_text=True)
        self.assertIn("请协助确认客户测试结论。", html)
        self.assertIn("@other@test.com", html)
        self.assertIn("发表留言", html)

    def test_comment_rejects_member_from_another_organization(self) -> None:
        project_id = self._seed_project()
        with app.app_context():
            outsider_id = self._add_user("outsider@test.com")
            other_org = Organization(name="其他组织")
            db.session.add(other_org)
            db.session.flush()
            outsider_membership = OrganizationMembership(
                organization_id=other_org.id, user_id=outsider_id
            )
            outsider_membership.set_roles(["sales"])
            db.session.add(outsider_membership)
            db.session.commit()
        self._login()
        response = self.client.post(
            f"/api/v1/customer-projects/projects/{project_id}/comments",
            json={"body": "不应跨组织提及", "mention_user_ids": [outsider_id]},
            headers={"Idempotency-Key": "invalid-comment-mention"},
        )
        self.assertEqual(response.status_code, 422)
        self.assertEqual(response.get_json()["error"]["code"], "VALIDATION_ERROR")

    def test_non_member_project_is_not_disclosed(self) -> None:
        project_id = self._seed_project()
        self._login("other@test.com")
        response = self.client.get(f"/api/v1/customer-projects/projects/{project_id}")
        self.assertEqual(response.status_code, 404)

    def test_materials_allow_zero_or_multiple_competitors(self) -> None:
        project_id = self._seed_project()
        self._login()
        first = self.client.post(
            f"/api/v1/customer-projects/projects/{project_id}/materials",
            json={"promoted_brand": "Mavis", "promoted_mpn": "MPX-8100", "category_code": "Power IC"},
            headers={"Idempotency-Key": "material-one"},
        )
        self.assertEqual(first.status_code, 201)
        material_id = first.get_json()["data"]["id"]
        material_retry = self.client.post(
            f"/api/v1/customer-projects/projects/{project_id}/materials",
            json={"promoted_brand": "Mavis", "promoted_mpn": "MPX-8100", "category_code": "Power IC"},
            headers={"Idempotency-Key": "material-one"},
        )
        self.assertEqual(material_retry.status_code, 201)
        self.assertEqual(material_retry.get_json()["data"]["id"], material_id)
        pending_payload = {"brand": "竞品品牌", "model_pending": True}
        pending = self.client.post(
            f"/api/v1/customer-projects/materials/{material_id}/competitors",
            json=pending_payload,
            headers={"Idempotency-Key": "competitor-one"},
        )
        self.assertEqual(pending.status_code, 201)
        competitor_retry = self.client.post(
            f"/api/v1/customer-projects/materials/{material_id}/competitors",
            json=pending_payload,
            headers={"Idempotency-Key": "competitor-one"},
        )
        self.assertEqual(competitor_retry.status_code, 201)
        self.assertEqual(competitor_retry.get_json()["data"]["id"], pending.get_json()["data"]["id"])
        second = self.client.post(
            f"/api/v1/customer-projects/materials/{material_id}/competitors",
            json={"brand": "另一品牌", "mpn": "ALT-100"},
            headers={"Idempotency-Key": "competitor-two"},
        )
        self.assertEqual(second.status_code, 201)
        with app.app_context():
            self.assertEqual(db.session.query(ProjectMaterial).count(), 1)
            self.assertEqual(db.session.query(MaterialCompetitor).count(), 2)

    def test_material_and_competitor_edit_conflict_and_soft_delete(self) -> None:
        project_id = self._seed_project()
        self._login()
        created_material = self.client.post(
            f"/api/v1/customer-projects/projects/{project_id}/materials",
            json={
                "promoted_brand": "Mavis",
                "promoted_mpn": "MPX-8100",
                "category_code": "Power IC",
            },
            headers={"Idempotency-Key": "editable-material"},
        )
        self.assertEqual(created_material.status_code, 201)
        material_id = created_material.get_json()["data"]["id"]

        updated_material = self.client.patch(
            f"/api/v1/customer-projects/materials/{material_id}",
            json={
                "promoted_brand": "Mavis Semi",
                "promoted_mpn": "MPX-8200",
                "category_code": "电源管理",
                "application_position": "域控制器电源",
                "machine_quantity": "4",
                "is_primary": True,
            },
            headers={"If-Match": '"1"'},
        )
        self.assertEqual(updated_material.status_code, 200)
        material_payload = updated_material.get_json()["data"]
        self.assertEqual(material_payload["promoted_brand"], "Mavis Semi")
        self.assertEqual(material_payload["machine_quantity"], "4")
        self.assertTrue(material_payload["is_primary"])
        self.assertEqual(material_payload["version"], 2)

        stale_material = self.client.patch(
            f"/api/v1/customer-projects/materials/{material_id}",
            json={"application_position": "不应覆盖"},
            headers={"If-Match": '"1"'},
        )
        self.assertEqual(stale_material.status_code, 409)
        self.assertEqual(
            stale_material.get_json()["error"]["code"], "MATERIAL_VERSION_CONFLICT"
        )
        self.assertEqual(
            stale_material.get_json()["error"]["details"]["current_version"], 2
        )

        created_competitor = self.client.post(
            f"/api/v1/customer-projects/materials/{material_id}/competitors",
            json={"brand": "竞品品牌", "model_pending": True},
            headers={"Idempotency-Key": "editable-competitor"},
        )
        self.assertEqual(created_competitor.status_code, 201)
        competitor_id = created_competitor.get_json()["data"]["id"]
        updated_competitor = self.client.patch(
            f"/api/v1/customer-projects/competitors/{competitor_id}",
            json={
                "brand": "竞品品牌二代",
                "mpn": "ALT-200",
                "model_pending": False,
                "distributor": "渠道 A",
                "quoted_price": "0.85",
                "observed_at": "2026-08-28",
            },
            headers={"If-Match": '"1"'},
        )
        self.assertEqual(updated_competitor.status_code, 200)
        competitor_payload = updated_competitor.get_json()["data"]
        self.assertEqual(competitor_payload["mpn"], "ALT-200")
        self.assertEqual(competitor_payload["quoted_price"], "0.85")
        self.assertEqual(competitor_payload["version"], 2)
        detail_before_delete = self.client.get(
            f"/customer-projects/projects/{project_id}"
        ).get_data(as_text=True)
        self.assertIn("编辑推广物料", detail_before_delete)
        self.assertIn("编辑竞争方案", detail_before_delete)

        stale_delete = self.client.delete(
            f"/api/v1/customer-projects/competitors/{competitor_id}",
            json={"reason": "信息已失效"},
            headers={"If-Match": '"1"'},
        )
        self.assertEqual(stale_delete.status_code, 409)
        deleted_competitor = self.client.delete(
            f"/api/v1/customer-projects/competitors/{competitor_id}",
            json={"reason": "信息已失效"},
            headers={"If-Match": '"2"'},
        )
        self.assertEqual(deleted_competitor.status_code, 204)
        deleted_material = self.client.delete(
            f"/api/v1/customer-projects/materials/{material_id}",
            json={"reason": "客户更换方案"},
            headers={"If-Match": '"2"'},
        )
        self.assertEqual(deleted_material.status_code, 204)

        detail = self.client.get(f"/customer-projects/projects/{project_id}")
        self.assertNotIn("Mavis Semi", detail.get_data(as_text=True))
        with app.app_context():
            self.assertIsNotNone(db.session.get(ProjectMaterial, material_id).deleted_at)
            self.assertIsNotNone(db.session.get(MaterialCompetitor, competitor_id).deleted_at)
            actions = {
                row.action
                for row in db.session.scalars(
                    db.select(AuditEvent).where(
                        AuditEvent.object_id.in_([material_id, competitor_id])
                    )
                )
            }
            self.assertIn("updated", actions)
            self.assertIn("deleted", actions)

    def test_legacy_material_without_mpn_can_be_edited(self) -> None:
        project_id = self._seed_project()
        self._login()
        created = self.client.post(
            f"/api/v1/customer-projects/projects/{project_id}/materials",
            json={
                "promoted_brand": "Mavis",
                "promoted_mpn": "LEGACY-100",
            },
            headers={"Idempotency-Key": "legacy-material"},
        )
        self.assertEqual(created.status_code, 201)
        material_id = created.get_json()["data"]["id"]
        with app.app_context():
            material = db.session.get(ProjectMaterial, material_id)
            material.promoted_mpn = None
            material.normalized_mpn = None
            material.mpn_pending = False
            db.session.commit()

        response = self.client.post(
            f"/customer-projects/materials/{material_id}/commercial",
            data={
                "material_version": "1",
                "promoted_brand": "Mavis Updated",
                "promoted_mpn": "",
                "category_code": "Power IC",
                "mpn_pending": "off",
                "is_primary": "on",
            },
        )
        self.assertEqual(response.status_code, 302)
        with app.app_context():
            material = db.session.get(ProjectMaterial, material_id)
            self.assertEqual(material.promoted_brand, "Mavis Updated")
            self.assertTrue(material.mpn_pending)

    def test_customer_grade_can_be_created_and_updated(self) -> None:
        self._login()
        created = self.client.post(
            "/customer-projects/customers",
            data={"name": "评级客户", "grade": "B"},
            follow_redirects=True,
        )
        self.assertIn("B", created.get_data(as_text=True))
        with app.app_context():
            customer = db.session.scalar(db.select(Customer).where(Customer.name == "评级客户"))
            customer_id = customer.id
            self.assertEqual(customer.grade, "B")
        updated = self.client.post(
            f"/customer-projects/customers/{customer_id}/grade",
            data={"grade": "A"},
        )
        self.assertEqual(updated.status_code, 302)
        with app.app_context():
            self.assertEqual(db.session.get(Customer, customer_id).grade, "A")

    def test_material_price_conversion_permissions_and_excel_export(self) -> None:
        project_id = self._seed_project()
        app._fx_cache = {
            "rates": {"USD": 1.0, "CNY": 7.2},
            "updated": "test-rate",
            "ts": time.time(),
        }
        self._login()
        priced = self.client.post(
            f"/api/v1/customer-projects/projects/{project_id}/materials",
            json={
                "promoted_brand": "Mavis",
                "promoted_mpn": "MPX-9000",
                "machine_quantity": "2",
                "unit_price": "1",
                "currency": "USD",
            },
            headers={"Idempotency-Key": "priced-material"},
        )
        self.assertEqual(priced.status_code, 201)
        payload = priced.get_json()["data"]
        self.assertEqual(payload["unit_price_usd"], "1")
        self.assertEqual(payload["unit_price_cny_tax_included"], "8.136")
        material_id = payload["id"]
        cny_priced = self.client.patch(
            f"/api/v1/customer-projects/materials/{material_id}",
            json={"machine_quantity": "2", "unit_price": "8.136", "currency": "CNY"},
            headers={"If-Match": '"1"'},
        )
        self.assertEqual(cny_priced.status_code, 200)
        cny_payload = cny_priced.get_json()["data"]
        self.assertEqual(cny_payload["unit_price_usd"], "1")
        self.assertEqual(cny_payload["unit_price_cny_tax_included"], "8.136")
        self.assertEqual(cny_payload["version"], 2)
        self.client.post("/logout")
        self._login("manager@test.com")
        export = self.client.get("/customer-projects/projects/export.xlsx")
        self.assertEqual(export.status_code, 200)
        self.assertEqual(
            export.mimetype,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        workbook = load_workbook(BytesIO(export.data), data_only=True)
        rows = list(workbook["客户项目台账"].iter_rows(values_only=True))
        self.assertIn("产品名称", rows[0])
        self.assertIn("含税人民币单价", rows[0])
        self.assertEqual(rows[1][4], "车载电源控制器")
        self.assertEqual(rows[1][17], 1)
        self.assertAlmostEqual(rows[1][18], 8.136, places=6)
        self.assertEqual(rows[1][21], 240000)

        with app.app_context():
            event = db.session.scalar(
                db.select(AuditEvent)
                .where(
                    AuditEvent.object_type == "customer_project_export",
                    AuditEvent.action == "exported",
                )
                .order_by(AuditEvent.occurred_at.desc())
            )
            diff = json.loads(event.safe_diff_json)
            self.assertEqual(diff["file_sha256"], hashlib.sha256(export.data).hexdigest())
            self.assertEqual(diff["project_count"], 1)
            self.assertTrue(diff["includes_prices"])

        self.client.post("/logout")
        with app.app_context():
            db.session.add(
                ProjectMember(
                    organization_id=self.org_id,
                    project_id=project_id,
                    user_id=self.fae_id,
                    role_code="fae",
                )
            )
            db.session.commit()
        self._login("fae@test.com")
        quantity_only = self.client.patch(
            f"/api/v1/customer-projects/materials/{material_id}",
            json={"machine_quantity": "3"},
            headers={"If-Match": '"2"'},
        )
        self.assertEqual(quantity_only.status_code, 200)
        self.assertEqual(quantity_only.get_json()["data"]["machine_quantity"], "3")
        denied = self.client.patch(
            f"/api/v1/customer-projects/materials/{material_id}",
            json={"machine_quantity": "3", "unit_price": "2", "currency": "USD"},
            headers={"If-Match": '"3"'},
        )
        self.assertEqual(denied.status_code, 403)

    def test_price_form_accepts_decimal_values_and_renders_decimal_inputs(self) -> None:
        project_id = self._seed_project()
        app._fx_cache = {
            "rates": {"USD": 1.0, "CNY": 7.2},
            "updated": "test-rate",
            "ts": time.time(),
        }
        self._login()

        page = self.client.get(f"/customer-projects/projects/{project_id}")
        self.assertEqual(page.status_code, 200)
        html = page.get_data(as_text=True)
        self.assertIn(
            'name="unit_price" min="0" step="any"',
            html,
        )

        created = self.client.post(
            f"/customer-projects/projects/{project_id}/materials",
            data={
                "idempotency_key": "decimal-price-form",
                "promoted_brand": "Mavis",
                "promoted_mpn": "DECIMAL-PRICE",
                "machine_quantity": "2",
                "unit_price": "1.23456",
                "currency": "USD",
            },
        )
        self.assertEqual(created.status_code, 302)
        with app.app_context():
            material = db.session.scalar(
                db.select(ProjectMaterial).where(
                    ProjectMaterial.project_id == project_id,
                    ProjectMaterial.promoted_mpn == "DECIMAL-PRICE",
                )
            )
            self.assertIsNotNone(material)
            self.assertEqual(material.target_price, Decimal("1.23456"))
            self.assertEqual(material.unit_price_usd, Decimal("1.23456"))

    def test_material_quantity_is_integer_and_price_has_at_most_five_decimals(self) -> None:
        project_id = self._seed_project()
        app._fx_cache = {
            "rates": {"USD": 1.0, "CNY": 7.2},
            "updated": "test-rate",
            "ts": time.time(),
        }
        self._login()
        fractional_quantity = self.client.post(
            f"/api/v1/customer-projects/projects/{project_id}/materials",
            json={
                "promoted_brand": "Mavis",
                "promoted_mpn": "INTEGER-ONLY",
                "machine_quantity": "1.5",
            },
            headers={"Idempotency-Key": "fractional-machine-quantity"},
        )
        self.assertEqual(fractional_quantity.status_code, 422)
        self.assertEqual(
            fractional_quantity.get_json()["error"]["field_errors"]["machine_quantity"],
            "请输入整数",
        )
        excessive_price = self.client.post(
            f"/api/v1/customer-projects/projects/{project_id}/materials",
            json={
                "promoted_brand": "Mavis",
                "promoted_mpn": "PRICE-LIMIT",
                "machine_quantity": "2",
                "unit_price": "1.123456",
                "currency": "USD",
            },
            headers={"Idempotency-Key": "excessive-price"},
        )
        self.assertEqual(excessive_price.status_code, 422)
        self.assertEqual(
            excessive_price.get_json()["error"]["field_errors"]["unit_price"],
            "最多输入 5 位小数",
        )
        trailing_zero_price = self.client.post(
            f"/api/v1/customer-projects/projects/{project_id}/materials",
            json={
                "promoted_brand": "Mavis",
                "promoted_mpn": "PRICE-TRIM",
                "machine_quantity": "2",
                "unit_price": "1.230000",
                "currency": "USD",
            },
            headers={"Idempotency-Key": "trimmed-price"},
        )
        self.assertEqual(trailing_zero_price.status_code, 201)
        self.assertEqual(trailing_zero_price.get_json()["data"]["unit_price_usd"], "1.23")

    def test_controlled_export_policy_scopes_roles_prices_and_formula_text(self) -> None:
        project_id = self._seed_project()
        with app.app_context():
            project = db.session.get(CustomerProject, project_id)
            project.name = '=HYPERLINK("https://invalid.example","click")'
            db.session.commit()

        self._login("sales@test.com")
        listing = self.client.get("/customer-projects/projects")
        self.assertNotIn("/customer-projects/projects/export.xlsx", listing.get_data(as_text=True))
        self.assertEqual(
            self.client.get("/customer-projects/projects/export.xlsx").status_code,
            403,
        )

        self.client.post("/logout")
        self._login("manager@test.com")
        updated = self.client.post(
            "/admin/customer-projects/export-policy",
            data={
                "allowed_roles": "sales",
                "max_projects": "2000",
                "max_rows": "20000",
            },
        )
        self.assertEqual(updated.status_code, 302)
        self.client.post("/logout")
        self._login("sales@test.com")
        export = self.client.get("/customer-projects/projects/export.xlsx")
        self.assertEqual(export.status_code, 200)
        workbook = load_workbook(BytesIO(export.data), data_only=False)
        rows = list(workbook["客户项目台账"].iter_rows(values_only=True))
        self.assertNotIn("含税人民币单价", rows[0])
        self.assertTrue(rows[1][3].startswith("'="))
        with app.app_context():
            policy = db.session.scalar(db.select(ProjectExportPolicy))
            self.assertEqual(policy.allowed_roles, frozenset({"sales"}))
            self.assertFalse(policy.include_prices)

    def test_controlled_export_rejects_unbounded_result_and_audits_reason(self) -> None:
        self._seed_project()
        with app.app_context():
            membership = db.session.scalar(
                db.select(OrganizationMembership).where(
                    OrganizationMembership.user_id == self.sales_id
                )
            )
            customer = create_customer({"name": "第二客户"}, membership)
            db.session.flush()
            create_project(
                {
                    "customer_id": customer.id,
                    "name": "第二项目",
                    "product_name": "第二产品",
                    "annual_usage": "10",
                    "stage_code": "evaluation",
                    "primary_sales_user_id": self.sales_id,
                    "next_action": "继续跟进",
                    "next_follow_up_at": (
                        datetime.now(timezone.utc) + timedelta(days=2)
                    ).isoformat(),
                },
                membership,
                "second-project",
            )
            db.session.commit()

        self._login("manager@test.com")
        updated = self.client.post(
            "/admin/customer-projects/export-policy",
            data={
                "allowed_roles": ["organization_admin", "business_manager"],
                "include_prices": "on",
                "max_projects": "1",
                "max_rows": "1",
            },
        )
        self.assertEqual(updated.status_code, 302)
        blocked = self.client.get("/customer-projects/projects/export.xlsx")
        self.assertEqual(blocked.status_code, 302)
        with app.app_context():
            event = db.session.scalar(
                db.select(AuditEvent)
                .where(
                    AuditEvent.object_type == "customer_project_export",
                    AuditEvent.action == "blocked",
                )
                .order_by(AuditEvent.occurred_at.desc())
            )
            self.assertIsNotNone(event)
            diff = json.loads(event.safe_diff_json)
            self.assertEqual(diff["reason"], "project_limit_exceeded")
            self.assertEqual(diff["max_projects"], 1)

    def test_personal_and_organization_saved_views_are_scoped_and_audited(self) -> None:
        self._seed_project()
        self._login("sales@test.com")
        created = self.client.post(
            "/customer-projects/views",
            data={
                "name": "我的评估项目",
                "visibility": "personal",
                "q": "示例",
                "stage": "evaluation",
            },
        )
        self.assertEqual(created.status_code, 302)
        with app.app_context():
            personal = db.session.scalar(
                db.select(ProjectSavedView).where(ProjectSavedView.visibility == "personal")
            )
            personal_id = personal.id
            self.assertEqual(personal.filters, {"q": "示例", "stage": "evaluation"})
        active = self.client.get(f"/customer-projects/projects?view={personal_id}")
        self.assertEqual(active.status_code, 200)
        self.assertIn("我的评估项目", active.get_data(as_text=True))

        denied_publish = self.client.post(
            "/customer-projects/views",
            data={
                "name": "越权共享",
                "visibility": "organization",
                "q": "",
                "stage": "evaluation",
            },
        )
        self.assertEqual(denied_publish.status_code, 302)
        self.client.post("/logout")
        self._login("other@test.com")
        self.assertEqual(
            self.client.get(f"/customer-projects/projects?view={personal_id}").status_code,
            404,
        )

        self.client.post("/logout")
        self._login("manager@test.com")
        shared_response = self.client.post(
            "/customer-projects/views",
            data={
                "name": "组织重点项目",
                "visibility": "organization",
                "q": "",
                "stage": "evaluation",
            },
        )
        self.assertEqual(shared_response.status_code, 302)
        with app.app_context():
            shared = db.session.scalar(
                db.select(ProjectSavedView).where(
                    ProjectSavedView.visibility == "organization"
                )
            )
            shared_id = shared.id
            shared_version = shared.version

        self.client.post("/logout")
        self._login("other@test.com")
        self.assertEqual(
            self.client.get(f"/customer-projects/projects?view={shared_id}").status_code,
            200,
        )
        denied_delete = self.client.post(
            f"/customer-projects/views/{shared_id}/delete",
            data={"version": str(shared_version)},
        )
        self.assertEqual(denied_delete.status_code, 302)
        with app.app_context():
            self.assertIsNotNone(db.session.get(ProjectSavedView, shared_id))

        self.client.post("/logout")
        self._login("manager@test.com")
        deleted = self.client.post(
            f"/customer-projects/views/{shared_id}/delete",
            data={"version": str(shared_version)},
        )
        self.assertEqual(deleted.status_code, 302)
        with app.app_context():
            self.assertIsNone(db.session.get(ProjectSavedView, shared_id))
            actions = {
                row.action
                for row in db.session.scalars(
                    db.select(AuditEvent).where(
                        AuditEvent.object_type == "project_saved_view"
                    )
                )
            }
            self.assertEqual(actions, {"created", "deleted"})

    def test_stage_history_soft_delete_and_manager_restore(self) -> None:
        project_id = self._seed_project()
        self._login()
        transition = self.client.post(
            f"/api/v1/customer-projects/projects/{project_id}/stage-transitions",
            json={"to_stage_code": "initiated", "reason": "客户确认立项", "project_version": 1},
            headers={"Idempotency-Key": "stage-one"},
        )
        self.assertEqual(transition.status_code, 201)
        self.assertEqual(transition.get_json()["data"]["from_stage_code"], "evaluation")
        deleted = self.client.delete(
            f"/api/v1/customer-projects/projects/{project_id}",
            json={"reason": "测试软删除"},
            headers={"If-Match": '"2"'},
        )
        self.assertEqual(deleted.status_code, 204)
        self.assertEqual(self.client.get(f"/api/v1/customer-projects/projects/{project_id}").status_code, 404)
        self.client.post("/logout")
        self._login("manager@test.com")
        restored = self.client.post(
            f"/api/v1/customer-projects/trash/projects/{project_id}/restore"
        )
        self.assertEqual(restored.status_code, 200)
        self.assertEqual(restored.get_json()["data"]["version"], 4)

    def test_dashboard_and_detail_render_responsive_product_pages(self) -> None:
        project_id = self._seed_project()
        self._login()
        dashboard = self.client.get("/customer-projects/")
        self.assertEqual(dashboard.status_code, 200)
        self.assertIn("今天先推进什么", dashboard.get_data(as_text=True))
        home = self.client.get("/").get_data(as_text=True)
        self.assertIn("客户项目工作台", home)
        self.assertIn('href="/customer-projects/"', home)
        detail = self.client.get(f"/customer-projects/projects/{project_id}")
        self.assertEqual(detail.status_code, 200)
        html = detail.get_data(as_text=True)
        self.assertIn("新增跟进", html)
        self.assertIn("尚未添加推广物料", html)
        self.assertIn("编辑项目基础信息", html)
        self.assertIn("data-project-title", html)
        self.assertIn("data-project-edit", html)
        self.assertNotIn("cp-form-panel cp-sticky", html)

    def test_local_day_bounds_use_organization_timezone(self) -> None:
        start, end = local_day_bounds(
            datetime(2026, 8, 27, 18, 30, tzinfo=timezone.utc), "Asia/Shanghai"
        )
        self.assertEqual(start, datetime(2026, 8, 27, 16, 0, tzinfo=timezone.utc))
        self.assertEqual(end, datetime(2026, 8, 28, 16, 0, tzinfo=timezone.utc))

    def test_project_edit_page_updates_core_fields_with_optimistic_lock(self) -> None:
        project_id = self._seed_project()
        self._login()
        response = self.client.post(
            f"/customer-projects/projects/{project_id}/edit",
            data={
                "project_version": "1",
                "name": "车载控制器二期",
                "product_name": "车规控制器二代",
                "annual_usage": "250000",
                "assessment_grade": "A",
                "probability_band": "70",
                "next_action": "确认量产排期",
                "next_follow_up_at": "2026-09-10T09:30",
                "expected_design_win_at": "2026-10-01",
                "expected_mass_production_at": "2027-01-15",
            },
        )
        self.assertEqual(response.status_code, 302)
        with app.app_context():
            project = db.session.get(CustomerProject, project_id)
            self.assertEqual(project.name, "车载控制器二期")
            self.assertEqual(project.product_name, "车规控制器二代")
            self.assertEqual(str(project.annual_usage), "250000.0000")
            self.assertEqual(project.assessment_grade, "A")
            self.assertEqual(project.probability_band, 70)
            self.assertEqual(project.next_action, "确认量产排期")
            self.assertEqual(project.next_follow_up_at.replace(tzinfo=timezone.utc).hour, 1)
            self.assertEqual(project.expected_design_win_at.isoformat(), "2026-10-01")
            self.assertEqual(project.expected_mass_production_at.isoformat(), "2027-01-15")
            self.assertEqual(project.version, 2)

    def test_project_edit_rejects_invalid_probability_without_server_error(self) -> None:
        project_id = self._seed_project()
        self._login()
        response = self.client.patch(
            f"/api/v1/customer-projects/projects/{project_id}",
            json={"probability_band": "invalid"},
            headers={"If-Match": '"1"'},
        )
        self.assertEqual(response.status_code, 422)
        self.assertEqual(response.get_json()["error"]["code"], "VALIDATION_ERROR")

    def test_project_annual_usage_requires_positive_integer(self) -> None:
        project_id = self._seed_project()
        self._login()
        response = self.client.patch(
            f"/api/v1/customer-projects/projects/{project_id}",
            json={"annual_usage": "120000.5"},
            headers={"If-Match": '"1"'},
        )
        self.assertEqual(response.status_code, 422)
        self.assertEqual(
            response.get_json()["error"]["field_errors"]["annual_usage"],
            "请输入大于 0 的整数",
        )

    def test_material_opportunity_categories_drive_market_scope(self) -> None:
        project_id = self._seed_project()
        with app.app_context():
            project = db.session.get(CustomerProject, project_id)
            materials = [
                ProjectMaterial(
                    organization_id=self.org_id,
                    project_id=project_id,
                    opportunity_type="design_in",
                    promoted_brand="Mavis",
                    promoted_mpn="DESIGN-1",
                    machine_quantity=Decimal("2"),
                    unit_price_usd=Decimal("1"),
                    idempotency_key="scope-design",
                    created_by_user_id=self.sales_id,
                    updated_by_user_id=self.sales_id,
                ),
                ProjectMaterial(
                    organization_id=self.org_id,
                    project_id=project_id,
                    opportunity_type="design_win",
                    promoted_brand="Mavis",
                    promoted_mpn="WIN-1",
                    machine_quantity=Decimal("1"),
                    unit_price_usd=Decimal("0.5"),
                    idempotency_key="scope-design-win",
                    created_by_user_id=self.sales_id,
                    updated_by_user_id=self.sales_id,
                ),
                ProjectMaterial(
                    organization_id=self.org_id,
                    project_id=project_id,
                    opportunity_type="matched_opportunity",
                    promoted_brand="Mavis",
                    promoted_mpn="MATCH-1",
                    machine_quantity=Decimal("1"),
                    unit_price_usd=Decimal("0.5"),
                    idempotency_key="scope-matched",
                    created_by_user_id=self.sales_id,
                    updated_by_user_id=self.sales_id,
                ),
                ProjectMaterial(
                    organization_id=self.org_id,
                    project_id=project_id,
                    opportunity_type="competitive_opportunity",
                    promoted_brand="",
                    machine_quantity=Decimal("3"),
                    idempotency_key="scope-competitive",
                    created_by_user_id=self.sales_id,
                    updated_by_user_id=self.sales_id,
                ),
            ]
            db.session.add_all(materials)
            db.session.flush()
            competitors = [
                MaterialCompetitor(
                    organization_id=self.org_id,
                    project_material_id=materials[3].id,
                    brand="Rival",
                    mpn="RIV-1",
                    quoted_price=Decimal("0.1"),
                    idempotency_key="scope-competitor-low",
                    created_by_user_id=self.sales_id,
                    updated_by_user_id=self.sales_id,
                ),
                MaterialCompetitor(
                    organization_id=self.org_id,
                    project_material_id=materials[3].id,
                    brand="Rival",
                    mpn="RIV-2",
                    quoted_price=Decimal("0.25"),
                    idempotency_key="scope-competitor-high",
                    created_by_user_id=self.sales_id,
                    updated_by_user_id=self.sales_id,
                ),
            ]
            db.session.add_all(competitors)
            db.session.commit()
            competitors_by_material = {materials[3].id: competitors}
            scope = build_market_scope(project, materials, competitors_by_material)
            self.assertEqual(scope["tam_usd"], Decimal("450000.00"))
            self.assertEqual(scope["sam_usd"], Decimal("360000.00"))
            self.assertEqual(scope["som_usd"], Decimal("300000.00"))
        self._login()
        page = self.client.get(f"/customer-projects/projects/{project_id}")
        html = page.get_data(as_text=True)
        self.assertIn("USD 450,000.00", html)
        self.assertIn('data-opportunity-type="competitive_opportunity"', html)
        self.assertIn('data-opportunity-type="design_win"', html)

    def test_lost_material_records_competitor_info_only(self) -> None:
        project_id = self._seed_project()
        self._login()
        created = self.client.post(
            f"/api/v1/customer-projects/projects/{project_id}/materials",
            json={
                "opportunity_type": "competitive_opportunity",
                "machine_quantity": "3",
                "category_code": "Power IC",
            },
            headers={"Idempotency-Key": "lost-material"},
        )
        self.assertEqual(created.status_code, 201)
        material_id = created.get_json()["data"]["id"]
        self.assertEqual(created.get_json()["data"]["promoted_brand"], "")
        competitor = self.client.post(
            f"/api/v1/customer-projects/materials/{material_id}/competitors",
            json={"brand": "Rival", "mpn": "RIV-9", "quoted_price": "0.4"},
            headers={"Idempotency-Key": "lost-competitor"},
        )
        self.assertEqual(competitor.status_code, 201)
        # Lost 物料的年度机会金额按竞品报价估算：120000 × 3 × 0.4
        page = self.client.get(f"/customer-projects/projects/{project_id}")
        html = page.get_data(as_text=True)
        self.assertIn('data-annual-value="144000.00"', html)

    def test_lost_transition_requires_promoted_material_info(self) -> None:
        project_id = self._seed_project()
        self._login()
        created = self.client.post(
            f"/api/v1/customer-projects/projects/{project_id}/materials",
            json={"opportunity_type": "competitive_opportunity"},
            headers={"Idempotency-Key": "lost-to-design-in"},
        )
        self.assertEqual(created.status_code, 201)
        material_id = created.get_json()["data"]["id"]
        missing_info = self.client.patch(
            f"/api/v1/customer-projects/materials/{material_id}",
            json={"opportunity_type": "design_in"},
            headers={"If-Match": '"1"'},
        )
        self.assertEqual(missing_info.status_code, 422)
        self.assertIn("Lost", missing_info.get_json()["error"]["message"])
        completed = self.client.patch(
            f"/api/v1/customer-projects/materials/{material_id}",
            json={
                "opportunity_type": "design_in",
                "promoted_brand": "Mavis",
                "promoted_mpn": "MPX-9000",
            },
            headers={"If-Match": '"1"'},
        )
        self.assertEqual(completed.status_code, 200)
        self.assertEqual(completed.get_json()["data"]["opportunity_type"], "design_in")

    def test_opportunity_badge_quick_switch_preserves_flags(self) -> None:
        """徽章快速转换表单（最小字段集）应保留主推/待确认状态并完成类型转换。"""
        project_id = self._seed_project()
        self._login()
        created = self.client.post(
            f"/api/v1/customer-projects/projects/{project_id}/materials",
            json={
                "opportunity_type": "design_in",
                "promoted_brand": "Mavis",
                "promoted_mpn": "QUICK-1",
                "is_primary": True,
            },
            headers={"Idempotency-Key": "badge-switch"},
        )
        self.assertEqual(created.status_code, 201)
        material_id = created.get_json()["data"]["id"]
        # 模拟徽章转换：仅提交类型 + 版本 + 回填 is_primary
        page = self.client.post(
            f"/customer-projects/materials/{material_id}/commercial",
            data={
                "opportunity_type": "matched_opportunity",
                "material_version": "1",
                "is_primary": "on",
            },
            follow_redirects=True,
        )
        self.assertEqual(page.status_code, 200)
        html = page.get_data(as_text=True)
        self.assertIn('data-opportunity-type="matched_opportunity"', html)
        self.assertIn('data-is-primary="1"', html)
        # Lost 转出（最小字段集、无品牌）应被服务端拒绝
        to_lost = self.client.post(
            f"/customer-projects/materials/{material_id}/commercial",
            data={
                "opportunity_type": "competitive_opportunity",
                "material_version": "2",
                "is_primary": "on",
            },
            follow_redirects=True,
        )
        self.assertEqual(to_lost.status_code, 200)
        back = self.client.post(
            f"/customer-projects/materials/{material_id}/commercial",
            data={"opportunity_type": "design_in", "material_version": "3"},
            follow_redirects=True,
        )
        self.assertEqual(back.status_code, 200)
        self.assertIn("Lost", back.get_data(as_text=True))

    def test_manager_reactivates_terminal_project_without_losing_history(self) -> None:
        project_id = self._seed_project()
        self._login("manager@test.com")
        closed = self.client.post(
            f"/api/v1/customer-projects/projects/{project_id}/stage-transitions",
            json={
                "to_stage_code": "lost",
                "reason": "客户原方案取消",
                "close_reason_code": "customer_cancelled",
                "close_notes": "预算调整，保留后续机会。",
                "project_version": 1,
            },
            headers={"Idempotency-Key": "close-before-reactivation"},
        )
        self.assertEqual(closed.status_code, 201)
        payload = {
            "to_stage_code": "sampling",
            "reason": "客户新平台恢复验证",
            "next_action": "安排新样品",
            "next_follow_up_at": (datetime.now(timezone.utc) + timedelta(days=3)).isoformat(),
            "primary_sales_user_id": self.sales_id,
            "project_version": 2,
        }
        reopened = self.client.post(
            f"/api/v1/customer-projects/projects/{project_id}/reactivate",
            json=payload,
            headers={"Idempotency-Key": "reactivate-one"},
        )
        self.assertEqual(reopened.status_code, 201)
        retry = self.client.post(
            f"/api/v1/customer-projects/projects/{project_id}/reactivate",
            json=payload,
            headers={"Idempotency-Key": "reactivate-one"},
        )
        self.assertEqual(retry.status_code, 201)
        with app.app_context():
            project = db.session.get(CustomerProject, project_id)
            self.assertEqual(project.stage_code, "sampling")
            self.assertEqual(project.version, 3)
            self.assertEqual(project.close_reason_code, "customer_cancelled")
            events = list(
                db.session.scalars(
                    db.select(ProjectStageEvent)
                    .where(ProjectStageEvent.project_id == project_id)
                    .order_by(ProjectStageEvent.occurred_at)
                )
            )
            self.assertEqual([item.to_stage_code for item in events], ["evaluation", "lost", "sampling"])

    def test_derived_project_copies_selected_assets_but_not_activity_history(self) -> None:
        project_id = self._seed_project()
        self._login()
        activity = self.client.post(
            f"/api/v1/customer-projects/projects/{project_id}/activities",
            json={
                "activity_type": "meeting",
                "summary": "来源项目跟进",
                "next_action": "等待结论",
                "next_follow_up_at": (datetime.now(timezone.utc) + timedelta(days=2)).isoformat(),
                "project_version": 1,
            },
            headers={"Idempotency-Key": "source-activity"},
        )
        self.assertEqual(activity.status_code, 201)
        material = self.client.post(
            f"/api/v1/customer-projects/projects/{project_id}/materials",
            json={"promoted_brand": "Mavis", "promoted_mpn": "MPX-100"},
            headers={"Idempotency-Key": "source-material"},
        )
        material_id = material.get_json()["data"]["id"]
        competitor = self.client.post(
            f"/api/v1/customer-projects/materials/{material_id}/competitors",
            json={"brand": "竞品甲", "mpn": "ALT-1", "distributor": "渠道甲"},
            headers={"Idempotency-Key": "source-competitor"},
        )
        self.assertEqual(competitor.status_code, 201)
        payload = {
            "name": "车载控制器二代",
            "product_name": "车载电源控制器二代",
            "annual_usage": "180000",
            "stage_code": "evaluation",
            "next_action": "确认二代规格",
            "next_follow_up_at": (datetime.now(timezone.utc) + timedelta(days=5)).isoformat(),
            "copy_members": True,
            "copy_materials": True,
            "copy_competitors": True,
        }
        created = self.client.post(
            f"/api/v1/customer-projects/projects/{project_id}/derive",
            json=payload,
            headers={"Idempotency-Key": "derive-one"},
        )
        self.assertEqual(created.status_code, 201)
        derived_id = created.get_json()["data"]["id"]
        retry = self.client.post(
            f"/api/v1/customer-projects/projects/{project_id}/derive",
            json=payload,
            headers={"Idempotency-Key": "derive-one"},
        )
        self.assertEqual(retry.get_json()["data"]["id"], derived_id)
        with app.app_context():
            derived = db.session.get(CustomerProject, derived_id)
            self.assertEqual(derived.derived_from_project_id, project_id)
            self.assertNotEqual(derived.project_code, db.session.get(CustomerProject, project_id).project_code)
            self.assertEqual(
                db.session.scalar(db.select(db.func.count()).select_from(ProjectActivity).where(ProjectActivity.project_id == derived_id)),
                0,
            )
            copied_materials = list(db.session.scalars(db.select(ProjectMaterial).where(ProjectMaterial.project_id == derived_id)))
            self.assertEqual(len(copied_materials), 1)
            self.assertEqual(
                db.session.scalar(db.select(db.func.count()).select_from(MaterialCompetitor).where(MaterialCompetitor.project_material_id == copied_materials[0].id)),
                1,
            )

    def test_lifecycle_report_exposes_definition_and_respects_scope(self) -> None:
        project_id = self._seed_project()
        self._login("manager@test.com")
        material = self.client.post(
            f"/api/v1/customer-projects/projects/{project_id}/materials",
            json={
                "promoted_brand": "Mavis",
                "promoted_mpn": "MPX-RPT",
                "category_code": "Power IC",
            },
            headers={"Idempotency-Key": "report-material"},
        )
        material_id = material.get_json()["data"]["id"]
        self.client.post(
            f"/api/v1/customer-projects/materials/{material_id}/competitors",
            json={"brand": "竞品乙", "mpn": "ALT-RPT", "distributor": "渠道乙"},
            headers={"Idempotency-Key": "report-competitor"},
        )
        closed = self.client.post(
            f"/api/v1/customer-projects/projects/{project_id}/stage-transitions",
            json={
                "to_stage_code": "lost",
                "reason": "竞品锁定",
                "close_reason_code": "competition",
                "close_notes": "复盘完成",
                "project_version": 1,
            },
            headers={"Idempotency-Key": "report-close"},
        )
        self.assertEqual(closed.status_code, 201)
        response = self.client.get("/api/v1/customer-projects/reports/lifecycle?stage=lost")
        self.assertEqual(response.status_code, 200)
        report = response.get_json()["data"]
        self.assertEqual(report["summary"]["by_stage"]["lost"], 1)
        self.assertEqual(report["summary"]["lost_by_reason"]["competition"], 1)
        self.assertIn("不是阶段转化率", report["metadata"]["definition"])
        filtered = self.client.get(
            "/api/v1/customer-projects/reports/lifecycle?category=Power%20IC&competitor_brand=竞品乙&material_brand=Mavis"
        )
        self.assertEqual(filtered.status_code, 200)
        self.assertEqual(filtered.get_json()["data"]["summary"]["total"], 1)
        page = self.client.get("/customer-projects/lifecycle")
        self.assertEqual(page.status_code, 200)
        self.assertIn("项目生命周期汇总", page.get_data(as_text=True))


if __name__ == "__main__":
    unittest.main()
