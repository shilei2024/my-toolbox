"""
FCST merge tool.

Uploads one or more FCST Excel files (.xls / .xlsx), reads them line-by-line
(the original script reads every 4th row starting at row 2, E column = part
number, J–O columns = monthly forecast), aggregates forecasts by part number,
looks up the part number in the DB-backed PN mapping table for mfr_part /
brand, and produces a summary .xlsx with two sheets:

  1. 明细: 品号 / 原厂料号 / 品牌 / 每月预测
  2. 型号汇总: 按原厂料号再汇总

Privacy: uploaded source filenames are never surfaced. The result file is
named with a uuid so it carries no trace of the original file names, and
the output only contains part numbers / mfr parts / brands / months — no
customer names.
"""
from __future__ import annotations

import io
import logging
from collections import defaultdict
from pathlib import Path
from typing import Any

import openpyxl
import xlrd
from flask import (
    Blueprint,
    abort,
    current_app,
    jsonify,
    render_template,
    request,
    send_file,
    send_from_directory,
    url_for,
)
from werkzeug.datastructures import FileStorage

from auth.decorators import commit_usage, ensure_anon_id, remaining_for, require_usage
from extensions import db, limiter
from models import PnMapping
from utils.helpers import is_allowed_ext, safe_filename

logger = logging.getLogger(__name__)
tool_bp = Blueprint("fcst_merge", __name__)


# ===========================================================================
# Owner resolution — anonymous data is temporary, user data is permanent.
# ===========================================================================
def _current_owner() -> tuple[str, str]:
    """Return (owner_type, owner_id) for the current request.

    - Logged-in user → ('user', str(user.id)) — permanently saved.
    - Anonymous → ('anon', anon_id from session) — temporary, cleaned up
      after 24h of inactivity.
    """
    from flask_login import current_user
    if current_user.is_authenticated:
        return ("user", str(current_user.id))
    return ("anon", ensure_anon_id())


def _owner_filter():
    ot, oid = _current_owner()
    return db.and_(PnMapping.owner_type == ot, PnMapping.owner_id == oid)


# Month header mapping: last 2 chars of header like "2024-01" → 一月
_MONTH_MAPPING = {
    "01": "一月", "02": "二月", "03": "三月", "04": "四月",
    "05": "五月", "06": "六月", "07": "七月", "08": "八月",
    "09": "九月", "10": "十月", "11": "十一月", "12": "十二月",
}

# Columns we read (1-indexed, matching the original script):
#   E=5 (part number), J=10..O=15 (monthly forecasts)
_PN_COL = 5
_MONTH_COLS = [10, 11, 12, 13, 14, 15]


# ===========================================================================
# Page
# ===========================================================================
@tool_bp.get("/")
def index():
    from flask_login import current_user
    return render_template(
        "tools_base.html",
        tool={
            "id": "fcst_merge",
            "name": "FCST 预测合并",
            "icon": "bi-table",
            "color": "#0dcaf0",
        },
        remaining=remaining_for("fcst_merge"),
        body_template="tools/fcst_merge/_body.html",
        tool_js_list=["js/fcst_merge.js", "js/result.js"],
        is_anon=not current_user.is_authenticated,
    )


# ===========================================================================
# PN mapping CRUD
# ===========================================================================
@tool_bp.get("/api/pn")
def pn_list():
    """List PN mappings with optional search & pagination (owner-scoped)."""
    q = (request.args.get("q") or "").strip()
    page = max(1, request.args.get("page", 1, type=int))
    per_page = min(100, max(10, request.args.get("per_page", 20, type=int)))

    query = db.session.query(PnMapping).filter(_owner_filter())
    if q:
        like = f"%{q}%"
        query = query.filter(
            db.or_(
                PnMapping.part_number.like(like),
                PnMapping.mfr_part.like(like),
                PnMapping.brand.like(like),
            )
        )
    query = query.order_by(PnMapping.part_number.asc())
    total = query.count()
    rows = query.offset((page - 1) * per_page).limit(per_page).all()
    return jsonify(
        ok=True,
        items=[r.to_dict() for r in rows],
        total=total,
        page=page,
        per_page=per_page,
        pages=(total + per_page - 1) // per_page,
    )


@tool_bp.post("/api/pn")
def pn_create():
    data = request.get_json(silent=True) or {}
    pn = (data.get("part_number") or "").strip()
    if not pn:
        return jsonify(error="品号不能为空"), 400
    ot, oid = _current_owner()
    if (
        db.session.query(PnMapping)
        .filter(_owner_filter(), PnMapping.part_number == pn)
        .first()
    ):
        return jsonify(error=f"品号 {pn} 已存在"), 409
    row = PnMapping(
        part_number=pn,
        mfr_part=(data.get("mfr_part") or "").strip(),
        brand=(data.get("brand") or "").strip(),
        owner_type=ot,
        owner_id=oid,
    )
    db.session.add(row)
    db.session.commit()
    logger.info("pn_mapping: created %s for %s/%s", pn, ot, oid)
    return jsonify(ok=True, item=row.to_dict()), 201


@tool_bp.put("/api/pn/<int:row_id>")
def pn_update(row_id: int):
    data = request.get_json(silent=True) or {}
    row = db.session.get(PnMapping, row_id)
    if row is None:
        return jsonify(error="记录不存在"), 404
    # owner check: can only edit own rows
    ot, oid = _current_owner()
    if row.owner_type != ot or row.owner_id != oid:
        return jsonify(error="无权修改此记录"), 403
    new_pn = (data.get("part_number") or "").strip()
    if new_pn and new_pn != row.part_number:
        if (
            db.session.query(PnMapping)
            .filter(_owner_filter(), PnMapping.part_number == new_pn)
            .first()
        ):
            return jsonify(error=f"品号 {new_pn} 已存在"), 409
        row.part_number = new_pn
    if "mfr_part" in data:
        row.mfr_part = (data["mfr_part"] or "").strip()
    if "brand" in data:
        row.brand = (data["brand"] or "").strip()
    db.session.commit()
    return jsonify(ok=True, item=row.to_dict())


@tool_bp.delete("/api/pn/<int:row_id>")
def pn_delete(row_id: int):
    row = db.session.get(PnMapping, row_id)
    if row is None:
        return jsonify(error="记录不存在"), 404
    ot, oid = _current_owner()
    if row.owner_type != ot or row.owner_id != oid:
        return jsonify(error="无权删除此记录"), 403
    db.session.delete(row)
    db.session.commit()
    return jsonify(ok=True)


@tool_bp.post("/api/pn/import")
def pn_import():
    """Batch import from an .xlsx file with columns: 品号 / 原厂料号 / 品牌."""
    f = request.files.get("file")
    if f is None or not f.filename:
        return jsonify(error="请选择一个 Excel 文件。"), 400
    if not is_allowed_ext(f.filename, {"xlsx"}):
        return jsonify(error="仅支持 .xlsx 文件。"), 400

    try:
        # openpyxl read_only needs a seekable stream; FileStorage's stream may
        # be a SpooledTemporaryFile which isn't always seekable in the way
        # openpyxl wants. Read into a BytesIO to be safe.
        raw = f.read()
        wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True, read_only=True)
    except Exception as exc:  # noqa: BLE001
        return jsonify(error=f"解析失败：{exc}"), 400
    ws = wb.active

    added = 0
    updated = 0
    skipped = 0
    ot, oid = _current_owner()
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:  # header
            continue
        if not row or not row[0]:
            skipped += 1
            continue
        pn = str(row[0]).strip()
        mfr = str(row[1]).strip() if len(row) > 1 and row[1] else ""
        brand = str(row[2]).strip() if len(row) > 2 and row[2] else ""
        existing = (
            db.session.query(PnMapping)
            .filter(_owner_filter(), PnMapping.part_number == pn)
            .first()
        )
        if existing:
            existing.mfr_part = mfr
            existing.brand = brand
            updated += 1
        else:
            db.session.add(PnMapping(
                part_number=pn, mfr_part=mfr, brand=brand,
                owner_type=ot, owner_id=oid,
            ))
            added += 1
    db.session.commit()
    return jsonify(ok=True, added=added, updated=updated, skipped=skipped, owner=ot)


@tool_bp.get("/api/pn/export")
def pn_export():
    """Export the current owner's PN mapping table as .xlsx."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "品号映射"
    ws.append(["品号", "原厂料号", "品牌"])
    for r in (
        db.session.query(PnMapping)
        .filter(_owner_filter())
        .order_by(PnMapping.part_number.asc())
        .all()
    ):
        ws.append([r.part_number, r.mfr_part, r.brand])
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return send_file(
        out,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="pn_mapping_export.xlsx",
    )


@tool_bp.get("/api/pn/stats")
def pn_stats():
    total = db.session.query(PnMapping).filter(_owner_filter()).count()
    brands = (
        db.session.query(PnMapping.brand, db.func.count(PnMapping.id))
        .filter(_owner_filter())
        .group_by(PnMapping.brand)
        .all()
    )
    ot, _ = _current_owner()
    return jsonify(ok=True, total=total, brands={b: c for b, c in brands if b}, owner=ot)


# ===========================================================================
# FCST processing
# ===========================================================================
@tool_bp.post("/process")
@limiter.limit(lambda: current_app.config["RATELIMIT_TOOL"])
@require_usage("fcst_merge")
def process():
    is_ajax = request.accept_mimetypes.best == "application/json" or request.is_json
    files = request.files.getlist("files")
    if not files or all(not f.filename for f in files):
        return _fail("请至少上传一个 FCST 文件。", is_ajax)

    # 1. Read every uploaded file into a unified list of (part_number, {month: qty})
    #    keyed by part_number. Source filenames are deliberately not preserved.
    fcst_data: dict[str, dict[str, int]] = defaultdict(lambda: defaultdict(int))
    file_count = 0
    errors: list[str] = []

    for f in files:
        if not f or not f.filename:
            continue
        if not is_allowed_ext(f.filename, {"xls", "xlsx"}):
            errors.append("仅支持 .xls / .xlsx 文件，已跳过不支持的文件。")
            continue
        try:
            if f.filename.lower().endswith(".xls"):
                _read_xls_fcst(f, fcst_data)
            else:
                _read_xlsx_fcst(f, fcst_data)
            file_count += 1
        except Exception as exc:  # noqa: BLE001
            logger.warning("fcst_merge: failed to read a file: %s", exc)
            errors.append(f"某个文件解析失败：{exc}")

    if file_count == 0:
        return _fail("没有可处理的文件。" + (" " + "；".join(errors) if errors else ""), is_ajax)

    # 2. Build the PN lookup from DB (owner-scoped)
    pn_rows = db.session.query(PnMapping).filter(_owner_filter()).all()
    pn_map: dict[str, PnMapping] = {r.part_number: r for r in pn_rows}

    # 3. Resolve month order from the union of all months seen.
    month_order = ["一月", "二月", "三月", "四月", "五月", "六月",
                   "七月", "八月", "九月", "十月", "十一月", "十二月"]
    seen_months = set()
    for months in fcst_data.values():
        seen_months.update(months.keys())
    months_out = [m for m in month_order if m in seen_months]
    if not months_out:
        return _fail("没有读到任何月份预测数据，请检查文件格式。", is_ajax)

    # 4. Build detail rows: 品号 / 原厂料号 / 品牌 / 各月
    detail_rows = []
    unmatched_pns = 0
    for pn in sorted(fcst_data.keys()):
        m = pn_map.get(pn)
        if m:
            mfr = m.mfr_part
            brand = m.brand
        else:
            mfr = ""
            brand = ""
            unmatched_pn += 0  # noqa
            unmatched_pns += 1 if pn else 0
        row = [pn, mfr, brand] + [fcst_data[pn].get(mo, 0) for mo in months_out]
        detail_rows.append(row)

    # 5. Build summary by mfr_part
    summary: dict[str, dict[str, Any]] = defaultdict(lambda: {"brand": "", "months": defaultdict(int)})
    for row in detail_rows:
        mfr = row[1] or "(未映射)"
        summary[mfr]["brand"] = row[2]
        for i, mo in enumerate(months_out):
            summary[mfr]["months"][mo] += row[3 + i] if isinstance(row[3 + i], (int, float)) else 0

    # 6. Write the output workbook
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "明细"
    ws1.append(["品号", "原厂料号", "品牌"] + months_out)
    for row in detail_rows:
        ws1.append(row)

    ws2 = wb.create_sheet("型号汇总")
    ws2.append(["原厂料号", "品牌"] + months_out)
    for mfr in sorted(summary.keys()):
        info = summary[mfr]
        ws2.append([mfr, info["brand"]] + [info["months"].get(mo, 0) for mo in months_out])

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    data = out.getvalue()

    commit_usage("fcst_merge", success=True)
    download_name = safe_filename("fcst_merged.xlsx")

    if is_ajax:
        try:
            filename = _stage_to_uploads(download_name, data)
        except Exception as exc:  # noqa: BLE001
            logger.exception("fcst_merge: failed to stage output: %s", exc)
            return jsonify(error="保存结果失败，请稍后再试。"), 500
        return jsonify(
            ok=True,
            url=url_for("fcst_merge.download", filename=filename),
            filename=filename,
            size=len(data),
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            stats={
                "files_read": file_count,
                "part_numbers": len(fcst_data),
                "months": months_out,
                "unmatched_part_numbers": unmatched_pns,
                "errors": errors,
            },
        )

    return send_file(
        io.BytesIO(data),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=download_name,
    )


@tool_bp.get("/download/<path:filename>")
def download(filename: str):
    if not is_allowed_ext(filename, {"xlsx"}):
        abort(404)
    return send_from_directory(
        current_app.config["UPLOAD_DIR"],
        filename,
        as_attachment=True,
        download_name=filename,
    )


# ===========================================================================
# Readers
# ===========================================================================
def _month_from_header(value: Any) -> str | None:
    """Map a header like '2024-01' or '2024-1月' to '一月'."""
    if value is None:
        return None
    s = str(value).strip()
    if not s:
        return None
    # last 2 chars (handles '01'..'12'); fall back to stripping non-digits
    tail = s[-2:]
    if tail in _MONTH_MAPPING:
        return _MONTH_MAPPING[tail]
    # try to find a 2-digit month anywhere
    for i in range(len(s) - 1):
        seg = s[i:i + 2]
        if seg in _MONTH_MAPPING:
            return _MONTH_MAPPING[seg]
    return None


def _read_xlsx_fcst(f: FileStorage, fcst_data: dict) -> None:
    raw = f.read()
    wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True, read_only=True)
    for ws in wb.worksheets:
        _read_sheet_xlsx(ws, fcst_data)


def _read_sheet_xlsx(ws, fcst_data: dict) -> None:
    # Read everything into a list so we can peek at row 1 headers and step by 4.
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return
    # month headers from row 0 (Excel row 1), columns J..O
    month_headers: dict[int, str] = {}
    if rows[0]:
        for col in _MONTH_COLS:
            idx = col - 1
            if idx < len(rows[0]):
                mo = _month_from_header(rows[0][idx])
                if mo:
                    month_headers[col] = mo

    # iterate every 4th row starting at row index 1 (Excel row 2)
    for r in range(1, len(rows), 4):
        row = rows[r]
        if not row:
            continue
        if _PN_COL - 1 >= len(row):
            continue
        pn = row[_PN_COL - 1]
        if pn is None or str(pn).strip() == "":
            continue
        pn = str(pn).strip()
        for col, mo in month_headers.items():
            idx = col - 1
            if idx >= len(row):
                continue
            val = row[idx]
            if val is None:
                continue
            try:
                qty = int(float(val))
            except (TypeError, ValueError):
                continue
            fcst_data[pn][mo] += qty


def _read_xls_fcst(f: FileStorage, fcst_data: dict) -> None:
    # xlrd needs a file path or bytes; we save to a temp BytesIO-backed path.
    # xlrd 2.x only reads .xls (legacy BIFF format).
    import tempfile, os
    tmp_path = None
    try:
        with tempfile.NamedTemporaryFile(suffix=".xls", delete=False) as tmp:
            tmp.write(f.read())
            tmp_path = tmp.name
        book = xlrd.open_workbook(tmp_path)
        for sheet in book.sheets():
            _read_sheet_xls(sheet, fcst_data)
    finally:
        if tmp_path and os.path.exists(tmp_path):
            try:
                os.unlink(tmp_path)
            except OSError:
                pass


def _read_sheet_xls(sheet, fcst_data: dict) -> None:
    # month headers from row 0 (Excel row 1)
    month_headers: dict[int, str] = {}
    if sheet.ncols > 0:
        for col in _MONTH_COLS:
            if col - 1 < sheet.ncols:
                mo = _month_from_header(sheet.cell_value(0, col - 1))
                if mo:
                    month_headers[col] = mo

    # iterate every 4th row starting at row index 1 (Excel row 2)
    for r in range(1, sheet.nrows, 4):
        if _PN_COL - 1 >= sheet.ncols:
            continue
        pn = sheet.cell_value(r, _PN_COL - 1)
        if pn is None or str(pn).strip() == "":
            continue
        pn = str(pn).strip()
        for col, mo in month_headers.items():
            if col - 1 >= sheet.ncols:
                continue
            val = sheet.cell_value(r, col - 1)
            try:
                qty = int(float(val))
            except (TypeError, ValueError):
                continue
            fcst_data[pn][mo] += qty


# ===========================================================================
# helpers
# ===========================================================================
def _fail(message: str, is_ajax: bool = False):
    commit_usage("fcst_merge", success=False, message=message)
    if is_ajax:
        return jsonify(error=message), 400
    from flask import flash, redirect
    flash(message, "danger")
    return redirect(url_for("fcst_merge.index"))


def _stage_to_uploads(suggested_name: str, data: bytes) -> str:
    upload_dir: Path = current_app.config["UPLOAD_DIR"]
    upload_dir.mkdir(parents=True, exist_ok=True)
    target = upload_dir / suggested_name
    target.write_bytes(data)
    return suggested_name
