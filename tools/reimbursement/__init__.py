"""
报销助手 — 发票上传、OCR 识别、封面生成、批量打印。

流程：上传发票 → 识别/录入明细 → 填写报销信息 → 生成封面 → 打印
"""
from __future__ import annotations

import json
import os
import uuid
from datetime import datetime
from pathlib import Path

from flask import Blueprint, current_app, jsonify, render_template, request, session

from auth.decorators import commit_usage, remaining_for, require_usage
from extensions import csrf, db, limiter
from models import ReimbursementRecord

tool_bp = Blueprint("reimbursement", __name__)

# ---------------------------------------------------------------------------
# 参考数据 — 来自公司 ERP 系统的产品线/办事处/部门
# ---------------------------------------------------------------------------
PRODUCT_LINES = [
    {"name": "DIODES", "code": "01", "office": "深圳办"},
    {"name": "MSTAR", "code": "02", "office": "厦门办"},
    {"name": "MSTAR-OTHERS", "code": "0201", "office": "杭州办"},
    {"name": "MSTAR-GPS", "code": "0202", "office": "上海办"},
    {"name": "MSTAR-OTT", "code": "0203", "office": "北京办"},
    {"name": "MSTAR-SIGMA", "code": "0204", "office": "合肥办"},
    {"name": "星宸", "code": "0205", "office": "西安办"},
    {"name": "MSTAR-MTK", "code": "0206", "office": ""},
    {"name": "MSTAR-RAFAEL", "code": "0222", "office": ""},
    {"name": "3PEAK", "code": "03", "office": ""},
    {"name": "KEMET", "code": "04", "office": ""},
    {"name": "REALTEK", "code": "05", "office": ""},
    {"name": "PANASONIC", "code": "06", "office": ""},
    {"name": "SILERGY", "code": "07", "office": ""},
    {"name": "SILERGY AC-DC", "code": "0701", "office": ""},
    {"name": "SILERGY DC-DC", "code": "0702", "office": ""},
    {"name": "SILERGY OTHERS", "code": "0703", "office": ""},
    {"name": "VIC", "code": "08", "office": ""},
    {"name": "艾为", "code": "09", "office": ""},
    {"name": "VANCHIP", "code": "10", "office": ""},
    {"name": "LITEON", "code": "11", "office": ""},
    {"name": "长虹WIFI模块", "code": "12", "office": ""},
    {"name": "创达特", "code": "13", "office": ""},
    {"name": "MEGACHIPS", "code": "14", "office": ""},
    {"name": "兴芯微", "code": "15", "office": ""},
    {"name": "ZTE", "code": "16", "office": ""},
    {"name": "西安紫光国芯", "code": "17", "office": ""},
    {"name": "CHIP-LEAD", "code": "18", "office": ""},
    {"name": "VIVA", "code": "19", "office": ""},
    {"name": "SINOPOWER", "code": "20", "office": ""},
    {"name": "比亚迪", "code": "21", "office": ""},
    {"name": "芯天下", "code": "22", "office": ""},
    {"name": "东软载波", "code": "23", "office": ""},
    {"name": "FEELING-TECH", "code": "24", "office": ""},
    {"name": "明月微", "code": "25", "office": ""},
    {"name": "海矽美", "code": "26", "office": ""},
    {"name": "DREAMTECH", "code": "27", "office": ""},
    {"name": "新大陆", "code": "28", "office": ""},
    {"name": "SENSORTEK", "code": "29", "office": ""},
    {"name": "FITIPOWER", "code": "30", "office": ""},
    {"name": "信炜", "code": "31", "office": ""},
    {"name": "AMPHENOL", "code": "32", "office": ""},
    {"name": "P2I", "code": "33", "office": ""},
    {"name": "CHIPOWN", "code": "34", "office": ""},
    {"name": "ANYKA", "code": "35", "office": ""},
    {"name": "思泰迪", "code": "36", "office": ""},
    {"name": "I-PEX", "code": "37", "office": ""},
    {"name": "费用", "code": "38", "office": ""},
    {"name": "太盟", "code": "39", "office": ""},
    {"name": "CAPELLA", "code": "40", "office": ""},
    {"name": "IMAGIS", "code": "41", "office": ""},
    {"name": "希荻微", "code": "42", "office": ""},
    {"name": "EVEREST", "code": "43", "office": ""},
    {"name": "YMIN", "code": "44", "office": ""},
    {"name": "JOULWATT", "code": "45", "office": ""},
    {"name": "DSPG", "code": "46", "office": ""},
    {"name": "CME", "code": "47", "office": ""},
    {"name": "SIGMA-EZVIZ", "code": "48", "office": ""},
    {"name": "临泰微", "code": "49", "office": ""},
    {"name": "老威欣", "code": "50", "office": ""},
    {"name": "新威欣", "code": "51", "office": ""},
    {"name": "厦门威欣", "code": "52", "office": ""},
    {"name": "澎湃微", "code": "53", "office": ""},
    {"name": "TI", "code": "54", "office": ""},
    {"name": "UnitedSiC", "code": "55", "office": ""},
    {"name": "易冲", "code": "56", "office": ""},
    {"name": "海思", "code": "57", "office": ""},
    {"name": "停用存货", "code": "94", "office": ""},
    {"name": "外箱", "code": "98", "office": ""},
    {"name": "其他", "code": "99", "office": ""},
    {"name": "其他-SSFI", "code": "9901", "office": ""},
    {"name": "其他-新大陆", "code": "9902", "office": ""},
    {"name": "其他-KINETIC", "code": "9903", "office": ""},
    {"name": "其他-村田", "code": "9904", "office": ""},
    {"name": "其他-BELLING", "code": "9905", "office": ""},
    {"name": "其他-DAVICOM", "code": "9906", "office": ""},
    {"name": "其他-乐沪", "code": "9907", "office": ""},
    {"name": "其他-威富锐", "code": "9908", "office": ""},
    {"name": "其他-UTC", "code": "9909", "office": ""},
    {"name": "其他-美思先端", "code": "9910", "office": ""},
    {"name": "其他-珊口", "code": "9911", "office": ""},
    {"name": "其他-RUNIC", "code": "9912", "office": ""},
    {"name": "其他-欢创", "code": "9913", "office": ""},
    {"name": "其他-移柯", "code": "9914", "office": ""},
    {"name": "其他-博流", "code": "9915", "office": ""},
    {"name": "其他-SSM", "code": "9916", "office": ""},
    {"name": "其他-RichWave", "code": "9917", "office": ""},
    {"name": "其他-捷捷微", "code": "9918", "office": ""},
    {"name": "其他-炬佑", "code": "9920", "office": ""},
    {"name": "其他-启英泰伦", "code": "9921", "office": ""},
    {"name": "其他-众智", "code": "9922", "office": ""},
    {"name": "其他-金贝", "code": "9923", "office": ""},
    {"name": "其他-爱信诺", "code": "9924", "office": ""},
    {"name": "其他-齐展", "code": "9925", "office": ""},
    {"name": "其他-瀚昕微", "code": "9926", "office": ""},
    {"name": "其他-维晟", "code": "9927", "office": ""},
    {"name": "其他-致象尔微", "code": "9928", "office": ""},
    {"name": "其他-爱普特", "code": "9929", "office": ""},
    {"name": "其他-启达微", "code": "9930", "office": ""},
    {"name": "其他-芯朋微", "code": "9931", "office": ""},
    {"name": "其他-凌思微", "code": "9932", "office": ""},
    {"name": "其他-MOSFET", "code": "9933", "office": ""},
    {"name": "其他-国民技术", "code": "9934", "office": ""},
    {"name": "其他-零星", "code": "9999", "office": ""},
]

OFFICES = [
    "深圳办", "厦门办", "杭州办", "上海办", "北京办",
    "合肥办", "西安办", "贸易组", "",
]

DEPARTMENTS = [
    "业务部", "市场部", "行政部",
]

EXPENSE_CATEGORIES = [
    {"key": "entertainment", "label": "招待费", "default": 0},
    {"key": "travel_transport", "label": "出差交通费", "default": 0},
    {"key": "travel_hotel", "label": "出差住宿费", "default": 0},
    {"key": "local_transport", "label": "市内交通费", "default": 0},
    {"key": "vehicle", "label": "车辆费用", "default": 0},
    {"key": "communication", "label": "通讯费", "default": 0},
    {"key": "office_supplies", "label": "办公费", "default": 0},
    {"key": "delivery", "label": "快递费", "default": 0},
    {"key": "welfare", "label": "福利", "default": 0},
]

CUSTOMER_LEVELS = [
    {"name": "0-1", "label": "0-1"},
    {"name": "level 1", "label": "level 1"},
    {"name": "level 2", "label": "level 2"},
    {"name": "level 3", "label": "level 3"},
]

ENTERTAINMENT_CATEGORIES = ["餐费", "送礼", "其他"]


# ---------------------------------------------------------------------------
# 工具函数
# ---------------------------------------------------------------------------
def _safe_float(val, default: float = 0.0) -> float:
    """安全转换为 float，非数字返回 default。"""
    try:
        return float(val) if val not in (None, "", "-") else default
    except (ValueError, TypeError):
        return default


def _safe_filename(name: str) -> str:
    """过滤文件名中的非法字符。"""
    import re
    return re.sub(r'[\\/:*?"<>|]', '_', name)[:100]


# ---------------------------------------------------------------------------
# 金额大写转换
# ---------------------------------------------------------------------------
def _number_to_chinese(amount: float) -> str:
    """将金额转换为中文大写（如 5616.10 → 伍仟陆佰壹拾陆元壹角整）"""
    if amount == 0:
        return "零元整"
    digit_cn = ["零", "壹", "贰", "叁", "肆", "伍", "陆", "柒", "捌", "玖"]
    radices_cn = ["", "拾", "佰", "仟", "万", "拾", "佰", "仟", "亿"]
    fraction_cn = ["角", "分"]

    yuan = int(amount)
    # 用整数运算避免浮点精度问题：5616.10 → 561610 分
    total_fen = int(round(amount * 100))
    jiao = (total_fen // 10) % 10
    fen = total_fen % 10

    # 整数部分
    result = ""
    if yuan == 0:
        result = "零"
    else:
        s = str(yuan)
        n = len(s)
        need_zero = False
        for i, ch in enumerate(s):
            d = int(ch)
            pos = n - i - 1
            if d == 0:
                need_zero = True
                if pos % 4 == 0 and pos > 0:
                    result += radices_cn[pos]
                    need_zero = False
            else:
                if need_zero:
                    result += "零"
                    need_zero = False
                result += digit_cn[d] + radices_cn[pos]
        if result.endswith("零"):
            result = result[:-1]
    result += "元"

    # 小数部分
    if jiao == 0 and fen == 0:
        result += "整"
    else:
        if jiao > 0:
            result += digit_cn[jiao] + fraction_cn[0]
        if fen > 0:
            result += digit_cn[fen] + fraction_cn[1]
        else:
            result += "整"
    return result


# ---------------------------------------------------------------------------
# 路由
# ---------------------------------------------------------------------------
@tool_bp.get("/")
def index():
    return render_template(
        "tools_base.html",
        tool={
            "id": "reimbursement",
            "name": "报销助手",
            "icon": "bi-receipt",
            "color": "#198754",
        },
        remaining=remaining_for("reimbursement"),
        body_template="tools/reimbursement/_body.html",
    )


@tool_bp.post("/upload")
@csrf.exempt
@limiter.limit(lambda: "20/minute")
@require_usage("reimbursement")
def upload():
    """上传发票图片/PDF，返回缩略图 URL。"""
    f = request.files.get("file")
    if not f or f.filename == "":
        return jsonify(error="请选择文件"), 400

    ext = Path(f.filename).suffix.lower()
    allowed = {".jpg", ".jpeg", ".png", ".pdf", ".bmp", ".gif", ".webp"}
    if ext not in allowed:
        return jsonify(error=f"不支持的文件格式：{ext}"), 400

    # 保存到 uploads 目录
    upload_dir = Path(current_app.config["UPLOAD_DIR"]) / "reimbursement"
    upload_dir.mkdir(parents=True, exist_ok=True)

    file_id = uuid.uuid4().hex[:12]
    safe_name = f"{file_id}{ext}"
    filepath = upload_dir / safe_name
    f.save(str(filepath))

    # PDF → 生成缩略图 PNG + 高清预览图 + 返回 base64（Vercel 兼容）
    preview_filename = safe_name
    full_preview_filename = safe_name
    thumb_b64 = ""  # 缩略图 base64，前端直接显示不依赖服务器文件系统

    if ext == ".pdf":
        try:
            import base64 as b64
            import fitz
            doc = fitz.open(str(filepath))
            page = doc[0]

            zoom = 200 / page.rect.height
            mat = fitz.Matrix(zoom, zoom)
            pix = page.get_pixmap(matrix=mat, colorspace=fitz.csRGB)
            thumb_name = f"{file_id}_thumb.png"
            pix.save(str(upload_dir / thumb_name))
            preview_filename = thumb_name
            thumb_b64 = b64.b64encode(pix.tobytes("png")).decode("ascii")

            full_zoom = 1600 / page.rect.height
            full_mat = fitz.Matrix(full_zoom, full_zoom)
            full_pix = page.get_pixmap(matrix=full_mat, colorspace=fitz.csRGB)
            full_name = f"{file_id}_full.png"
            full_pix.save(str(upload_dir / full_name))
            full_preview_filename = full_name

            doc.close()
        except Exception:
            full_preview_filename = preview_filename
    else:
        # 非 PDF：生成小缩略图 base64
        try:
            import base64 as b64
            import io as _io
            from PIL import Image
            img = Image.open(str(filepath))
            w, h = img.size
            if h > 200:
                img = img.resize((int(w * 200 / h), 200), Image.LANCZOS)
            buf = _io.BytesIO()
            img.convert("RGB").save(buf, format="JPEG", quality=80)
            thumb_b64 = b64.b64encode(buf.getvalue()).decode("ascii")
        except Exception:
            pass

    commit_usage("reimbursement")

    return jsonify(
        success=True,
        file_id=file_id,
        filename=safe_name,
        original_name=f.filename,
        preview_url=f"/tools/reimbursement/preview/{preview_filename}",
        full_url=f"/tools/reimbursement/preview/{full_preview_filename}",
        thumb_b64=thumb_b64,
        size=filepath.stat().st_size,
    )


@tool_bp.get("/preview/<filename>")
def preview(filename):
    """返回上传文件的缩略图/预览。"""
    from flask import send_file

    # 安全校验：防止路径遍历
    if ".." in filename or "/" in filename or "\\" in filename:
        return jsonify(error="非法文件名"), 400

    filepath = Path(current_app.config["UPLOAD_DIR"]) / "reimbursement" / filename
    # 确保解析后的路径仍在上传目录内
    try:
        filepath.resolve().relative_to((Path(current_app.config["UPLOAD_DIR"]) / "reimbursement").resolve())
    except ValueError:
        return jsonify(error="非法路径"), 400

    if not filepath.exists():
        return jsonify(error="文件不存在"), 404

    ext = filepath.suffix.lower()
    mimetypes = {
        ".jpg": "image/jpeg",
        ".jpeg": "image/jpeg",
        ".png": "image/png",
        ".gif": "image/gif",
        ".webp": "image/webp",
        ".bmp": "image/bmp",
        ".pdf": "application/pdf",
    }
    return send_file(
        str(filepath),
        mimetype=mimetypes.get(ext, "application/octet-stream"),
        max_age=3600,
    )


@tool_bp.post("/ocr")
@csrf.exempt
@limiter.limit(lambda: "10/minute")
@require_usage("reimbursement")
def ocr():
    """
    OCR 识别发票信息。
    接收 JSON: {"file_id": "...", "image_base64": "..."}

    优先级：file_id（原始文件，质量最高）> image_base64（Vercel 兼容降级）
    OCR 策略：百度增值税发票识别 → 百度通用 OCR → PaddleOCR 本地识别 → 模拟降级
    """
    data = request.get_json(silent=True) or {}
    img_bytes = None

    # 方案 A：通过 file_id 读取服务器上的原始文件（质量最高，推荐）
    file_id = (data.get("file_id") or "").strip()
    if file_id:
        upload_dir = Path(current_app.config["UPLOAD_DIR"]) / "reimbursement"
        filepath = None
        for p in upload_dir.glob(f"{file_id}.*"):
            filepath = p
            break
        if filepath and filepath.exists():
            img_bytes = filepath.read_bytes()
            current_app.logger.info("OCR using original file: %s (%d bytes)", filepath.name, len(img_bytes))

    # 方案 B：base64 直接传入（Vercel 兼容，或前端降级）
    if not img_bytes:
        img_b64 = (data.get("image_base64") or "").strip()
        if img_b64:
            import base64
            # 移除可能的 data:xxx;base64, 前缀
            if "," in img_b64 and "base64" in img_b64:
                img_b64 = img_b64.split(",", 1)[1]
            try:
                img_bytes = base64.b64decode(img_b64)
                current_app.logger.info("OCR using base64: %d bytes", len(img_bytes))
            except Exception:
                return jsonify(error="base64 解码失败"), 400
        else:
            # 方案 C：文件上传（兼容旧版）
            f = request.files.get("file")
            if f and f.filename:
                img_bytes = f.read()
            else:
                return jsonify(error="请提供 file_id 或 image_base64"), 400

    if not img_bytes:
        return jsonify(error="图片数据为空"), 400

    # --- 策略 1: 百度云 OCR（优先，增值税发票识别 + 通用 OCR） ---
    baidu_key = os.environ.get("BAIDU_OCR_API_KEY", "").strip()
    baidu_secret = os.environ.get("BAIDU_OCR_SECRET_KEY", "").strip()

    if baidu_key and baidu_secret:
        try:
            result = _baidu_ocr_from_bytes(img_bytes, baidu_key, baidu_secret)
            if result:
                has_data = any(result.get(k) for k in ("invoice_number", "total_amount", "seller_name"))
                if has_data:
                    return jsonify(success=True, data=result, provider="baidu")
                else:
                    current_app.logger.warning("Baidu OCR returned empty fields, trying fallbacks...")
        except Exception as e:
            current_app.logger.warning("Baidu OCR error: %s", e)
    else:
        current_app.logger.info("Baidu OCR keys not configured, skipping...")

    # --- 策略 2: PaddleOCR 本地识别 ---
    try:
        # 先保存为临时文件（PaddleOCR 需要文件路径）
        upload_dir = Path(current_app.config["UPLOAD_DIR"]) / "reimbursement"
        tmp_path = upload_dir / f"_ocr_tmp_{uuid.uuid4().hex[:8]}.png"
        try:
            # 如果是 PDF 或大图片，先转为 PNG
            from PIL import Image
            import io as _io
            if img_bytes[:5] == b"%PDF-":
                import fitz
                doc = fitz.open(stream=img_bytes, filetype="pdf")
                page = doc[0]
                mat = fitz.Matrix(300 / 72, 300 / 72)  # 300 DPI for OCR
                pix = page.get_pixmap(matrix=mat, colorspace=fitz.csRGB)
                pix.save(str(tmp_path))
                doc.close()
            else:
                img = Image.open(_io.BytesIO(img_bytes))
                if img.mode in ("RGBA", "LA", "P"):
                    img = img.convert("RGB")
                w, h = img.size
                if max(w, h) > 4000:
                    ratio = 4000 / max(w, h)
                    img = img.resize((int(w * ratio), int(h * ratio)), Image.LANCZOS)
                img.save(str(tmp_path), format="PNG")

            result = _paddle_ocr_invoice(tmp_path)
            if result and (result.get("invoice_number") or result.get("total_amount")):
                return jsonify(success=True, data=result, provider="paddleocr")
        finally:
            if tmp_path.exists():
                try:
                    tmp_path.unlink()
                except Exception:
                    pass
    except Exception as e:
        current_app.logger.warning("PaddleOCR fallback error: %s", e)

    # --- 策略 3: 模拟降级 ---
    return jsonify(
        success=True,
        data={
            "invoice_number": "", "invoice_date": "",
            "seller_name": "", "amount_excluding_tax": "",
            "tax_amount": "", "total_amount": "", "description": "",
        },
        provider="mock",
        note="未配置 OCR 服务或识别失败，请手动填写。",
    )


# ---------------------------------------------------------------------------
# OCR 核心：从字节到结果（Vercel 兼容，无文件系统依赖）
# ---------------------------------------------------------------------------
def _baidu_ocr_from_bytes(img_bytes: bytes, api_key: str, secret_key: str) -> dict | None:
    """
    完整 OCR 流程（接收字节，无需文件路径）。
    PDF 字节 → PyMuPDF 转高分辨率图片 → 预处理多版本 → 增值税发票识别 → 通用 OCR 降级
    """
    import io as _io

    # Step 0: 检测是否为 PDF，是则转高分辨率图片
    img_list = []
    if img_bytes[:5] == b"%PDF-":
        try:
            import fitz
            doc = fitz.open(stream=img_bytes, filetype="pdf")
            for page_num in range(min(len(doc), 3)):
                page = doc[page_num]
                # 提高渲染 DPI 以获得更好的 OCR 效果
                mat = fitz.Matrix(300 / 72, 300 / 72)
                pix = page.get_pixmap(matrix=mat, colorspace=fitz.csRGB)
                jpg = pix.tobytes("jpeg")
                if len(jpg) > 2_500_000:
                    mat2 = fitz.Matrix(200 / 72, 200 / 72)
                    pix2 = page.get_pixmap(matrix=mat2, colorspace=fitz.csRGB)
                    jpg = pix2.tobytes("jpeg")
                img_list.append(jpg)
            doc.close()
        except Exception as e:
            current_app.logger.warning("PyMuPDF stream: %s", e)
            img_list = [img_bytes]
    else:
        # 非 PDF：大图片适度压缩，小图片保持原样
        if len(img_bytes) > 4_000_000:
            try:
                from PIL import Image
                img = Image.open(_io.BytesIO(img_bytes))
                w, h = img.size
                if max(w, h) > 3000:
                    img = img.resize((int(w * 3000 / max(w, h)), int(h * 3000 / max(w, h))), Image.LANCZOS)
                buf = _io.BytesIO()
                img.convert("RGB").save(buf, format="JPEG", quality=90)
                img_list = [buf.getvalue()]
            except:
                img_list = [img_bytes]
        else:
            img_list = [img_bytes]

    if not img_list:
        return None

    # Step 1: 获取 token
    try:
        resp = __import__("requests").get(
            "https://aip.baidubce.com/oauth/2.0/token",
            params={"grant_type": "client_credentials", "client_id": api_key, "client_secret": secret_key},
            timeout=10,
        )
        token = resp.json().get("access_token")
        if not token:
            current_app.logger.warning("Baidu token failed")
            return None
    except Exception as e:
        current_app.logger.warning("Baidu token: %s", e)
        return None

    # Step 2: 增值税发票识别（页×版本穷举）
    for raw in img_list:
        for ver in _preprocess_for_ocr(raw):
            words = _baidu_vat_call(ver, token)
            if words:
                r = _parse_vat(words)
                if r.get("invoice_number") or r.get("total_amount"):
                    return r

    # Step 3: 通用 OCR 降级
    best, best_cnt = None, 0
    for raw in img_list:
        for ver in _preprocess_for_ocr(raw):
            words_list = _baidu_gen_call(ver, token)
            if words_list and len(words_list) > 2:
                r = _parse_general(words_list)
                cnt = sum(1 for v in r.values() if v)
                if cnt > best_cnt:
                    best, best_cnt = r, cnt
                    if cnt >= 3:
                        return best
    return best


# ---------------------------------------------------------------------------
# 图片预处理
# ---------------------------------------------------------------------------
def _preprocess_for_ocr(img_bytes: bytes) -> list[bytes]:
    """生成多个预处理版本：原图 / 灰度增强 / 强对比度。"""
    import io as _io
    try:
        from PIL import Image, ImageEnhance, ImageFilter
    except ImportError:
        return [img_bytes]
    results = [img_bytes]
    try:
        img = Image.open(_io.BytesIO(img_bytes))
        w, h = img.size
        if max(w, h) > 3000:
            ratio = 3000 / max(w, h)
            img = img.resize((int(w * ratio), int(h * ratio)), Image.LANCZOS)
        if img.mode in ("RGBA", "P", "LA"):
            img = img.convert("RGB")
        gray = img.convert("L")
        enhancer = ImageEnhance.Contrast(gray)
        gray_enh = enhancer.enhance(2.0).filter(ImageFilter.SHARPEN)
        buf = _io.BytesIO(); gray_enh.save(buf, format="PNG"); results.append(buf.getvalue())
        strong = enhancer.enhance(3.0)
        buf2 = _io.BytesIO(); strong.save(buf2, format="PNG"); results.append(buf2.getvalue())
    except Exception:
        pass
    return results


# ---------------------------------------------------------------------------
# 百度 OCR 调用
# ---------------------------------------------------------------------------
def _baidu_vat_call(img_bytes: bytes, access_token: str) -> dict | None:
    """增值税发票识别 → words_result 或 None。"""
    import base64; from urllib.parse import quote
    b64 = base64.b64encode(img_bytes).decode("ascii")
    body = f"image={quote(b64, safe='')}".encode("utf-8")
    resp = __import__("requests").post(
        f"https://aip.baidubce.com/rest/2.0/ocr/v1/vat_invoice?access_token={access_token}",
        data=body, headers={"Content-Type": "application/x-www-form-urlencoded"}, timeout=20,
    )
    data = resp.json()
    if data.get("error_code"):
        current_app.logger.debug("Baidu VAT [%s]: %s", data["error_code"], data.get("error_msg", ""))
        return None
    return data.get("words_result")


def _baidu_gen_call(img_bytes: bytes, access_token: str) -> list | None:
    """通用 OCR → words_result 列表。"""
    import base64; from urllib.parse import quote
    b64 = base64.b64encode(img_bytes).decode("ascii")
    body = f"image={quote(b64, safe='')}".encode("utf-8")
    resp = __import__("requests").post(
        f"https://aip.baidubce.com/rest/2.0/ocr/v1/general_basic?access_token={access_token}",
        data=body, headers={"Content-Type": "application/x-www-form-urlencoded"}, timeout=15,
    )
    data = resp.json()
    return data.get("words_result") if not data.get("error_code") else None


# ---------------------------------------------------------------------------
# 结果解析
# ---------------------------------------------------------------------------
def _normalize_date(raw: str) -> str:
    """将各种中文/数字日期格式统一转为 YYYY-MM-DD。"""
    import re
    if not raw or not raw.strip():
        return ""
    raw = raw.strip()
    # 已经是 YYYY-MM-DD 或 YYYY/MM/DD
    m = re.search(r"(\d{4})[-/](\d{1,2})[-/](\d{1,2})", raw)
    if m:
        return f"{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3)):02d}"
    # 中文格式：2024年01月15日 / 2024年1月15日 / 2024年 01月 15日
    m = re.search(r"(\d{4})\s*年\s*(\d{1,2})\s*月\s*(\d{1,2})\s*日", raw)
    if m:
        return f"{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3)):02d}"
    # 纯数字：20240115
    m = re.search(r"(\d{4})(\d{2})(\d{2})", raw)
    if m and 8 <= len(raw.replace(" ", "")) <= 10:
        y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if 2000 <= y <= 2100 and 1 <= mo <= 12 and 1 <= d <= 31:
            return f"{y}-{mo:02d}-{d:02d}"
    return raw


def _parse_vat(words: dict) -> dict:
    """
    解析百度增值税发票识别结果。兼容全电发票和传统发票。
    """
    def _g(f):
        v = words.get(f, "")
        if isinstance(v, list) and v:
            item = v[0]
            # 全电发票返回 [{"row":"1","word":"*餐费"}] 格式
            if isinstance(item, dict) and "word" in item:
                return item["word"]
            return item if isinstance(item, str) else str(item)
        return str(v) if v and not isinstance(v, (list, dict)) else ""

    # 金额字段：全电发票 AmountInFiguers=价税合计, TotalAmount=不含税
    total = _g("AmountInFiguers") or _g("TotalAmount") or ""
    no_tax = _g("TotalAmount") or ""
    tax = _g("TaxAmount") or ""

    if not tax and total and no_tax:
        try:
            diff = round(float(total) - float(no_tax), 2)
            if diff > 0: tax = str(diff)
        except: pass
    if not tax and total:
        try:
            if not no_tax or abs(float(total) - float(no_tax)) < 0.005:
                no_tax = str(float(total)); tax = "0.00"
        except: pass
    if not total and no_tax: total = no_tax
    if not no_tax and total and tax:
        try: no_tax = str(round(float(total) - float(tax), 2))
        except: pass

    # 明细：全电发票 CommodityName 是 [{"row":"1","word":"*餐费"}]
    desc = _g("CommodityName") or _g("CommodityType") or ""
    if not desc:
        items = words.get("CommodityName", [])
        if isinstance(items, list) and items:
            parts = []
            for item in items:
                if isinstance(item, dict) and "word" in item:
                    parts.append(item["word"])
                elif isinstance(item, str):
                    parts.append(item)
            desc = "；".join(p.strip() for p in parts[:3] if p.strip())

    return {
        "invoice_number": _g("InvoiceNum") or _g("InvoiceCodeConfirm") or "",
        "invoice_date": _normalize_date(_g("InvoiceDate") or _g("InvoiceTime") or ""),
        "seller_name": _g("SellerName") or "",
        "amount_excluding_tax": no_tax,
        "tax_amount": tax,
        "total_amount": total,
        "description": str(desc or "")[:80],
    }


def _parse_general(words_list: list) -> dict:
    """用增强正则模式从 OCR 文本中提取发票字段。"""
    import re
    text = " ".join([w.get("words", "") for w in (words_list or [])])
    # 也准备一个无空格版本用于某些模式
    text_nospace = "".join([w.get("words", "") for w in (words_list or [])])

    def _first(patterns, source=text):
        for p in patterns:
            m = re.search(p, source)
            if m:
                # 取最后一个有值的捕获组
                for g in reversed(m.groups()):
                    if g is not None:
                        return g.replace(",", "").replace("，", "").strip()
        return ""

    def _first_amt(patterns, source=text):
        """专门匹配金额的模式。"""
        for p in patterns:
            m = re.search(p, source)
            if m:
                for g in reversed(m.groups()):
                    if g is not None:
                        val = g.replace(",", "").replace("，", "").strip()
                        # 验证是否为有效金额格式
                        if re.match(r'^\d+\.?\d*$', val):
                            return val
        return ""

    return {
        "invoice_number": _first([
            r"发票号码[：:\s]*([A-Za-z0-9\-]{8,30})",
            r"No[\.\s]*([A-Za-z0-9\-]{8,30})",
            r"号码[：:\s]*([A-Za-z0-9\-]{8,30})",
            r"发票代码[：:\s]*(\d{10,12})\s*[,\s]+\s*(\d{8,10})",  # 发票代码 + 发票号码
            r"(\d{20})",  # 全电发票 20 位号码
            r"(\d{10,12})\s+[¥￥]\s*\d",  # 10-12位数字后跟金额符号
            r"(\d{8,12})",  # 传统发票号码
        ]),
        "invoice_date": _normalize_date(_first([
            r"(\d{4}\s*年\s*\d{1,2}\s*月\s*\d{1,2}\s*日)",
            r"开票日期[：:\s]*(\d{4}[-/年]\d{1,2}[-/月]\d{1,2})",
            r"日期[：:\s]*(\d{4}[-/年]\d{1,2}[-/月]\d{1,2})",
            r"(\d{4}-\d{2}-\d{2})",
            r"(\d{4}/\d{2}/\d{2})",
            r"(\d{4}年\d{1,2}月\d{1,2}日)",
        ])),
        "seller_name": _first([
            r"[销銷]售方[名称稱][：:\s]*([^\s，\n,]{4,60}?(?:有限公司|股份有限公司|有限责任公司|科技|电子|实业|集团|工厂|中心|经营部|商店|事务所|工作室|商行|服务部)[^\s，\n]{0,20})",
            r"[销銷]售?[方务][：:\s]*([^\s，\n,]{4,60}?(?:公司|科技|工厂|中心|事务所)[^\s，\n]{0,20})",
            r"名称[：:\s]*([^\s，\n,]{4,60}?(?:有限公司|股份有限公司|科技|电子|实业|集团|经营部)[^\s，\n]{0,20})",
            r"销货单位[名称稱]?[：:\s]*([^\s，\n,]{4,60}?(?:公司|工厂|商店)[^\s，\n]{0,20})",
        ]),
        "amount_excluding_tax": _first_amt([
            r"金额[\(（]不含税[\)）]?\s*[¥￥]?\s*([\d,]+\.?\d{0,2})",
            r"不含税金额[：:\s]*[¥￥]?\s*([\d,]+\.?\d{0,2})",
            r"不含税价[：:\s]*[¥￥]?\s*([\d,]+\.?\d{0,2})",
        ]),
        "tax_amount": _first_amt([
            r"[税稅]额[：:\s]*[¥￥]?\s*([\d,]+\.?\d{0,2})",
            r"税[额費金][：:\s]*[¥￥]?\s*([\d,]+\.?\d{0,2})",
        ]),
        "total_amount": _first_amt([
            r"价税合计[\(（]?\s*[小写大写]?[\)）]?\s*[¥￥]\s*([\d,]+\.?\d{0,2})",
            r"价税[合計]计[\(（]?[小写]?[\)）]?\s*[¥￥]\s*([\d,]+\.?\d{0,2})",
            r"[合計]计[金]?\s*[¥￥]?\s*([\d,]+\.?\d{0,2})",
            r"小写[金额]?\s*[¥￥]?\s*([\d,]+\.\d{2})",
            r"[¥￥]\s*([\d,]+\.\d{2})",  # ¥符号后金额
            r"金额[合計]?[计]?\s*[¥￥]?\s*([\d,]+\.?\d{0,2})",
        ]),
        "description": _first([
            r"[货貨]物[或及].*?[名称稱][：:\s]*(.{1,60})",
            r"项目[名称稱][：:\s]*(.{1,60})",
            r"货物名称[：:\s]*(.{1,60})",
            r"[\\*\\*](\S{2,20})",  # *商品名* 格式
            r"[*＊](\S{2,20})[*＊]",  # *餐费* 等
        ])[:80],
    }


def _paddle_ocr_invoice(filepath: Path) -> dict | None:
    """本地 PaddleOCR 识别（需提前安装 paddleocr）。"""
    try:
        from paddleocr import PaddleOCR
    except ImportError:
        current_app.logger.warning("PaddleOCR not installed")
        return None

    try:
        ocr_engine = PaddleOCR(use_angle_cls=True, lang="ch", show_log=False)
        result = ocr_engine.ocr(str(filepath), cls=True)

        # 提取所有文本行
        lines = []
        if result and result[0]:
            for line_info in result[0]:
                text = line_info[1][0] if len(line_info) > 1 else ""
                conf = line_info[1][1] if len(line_info) > 1 else 0
                lines.append((text, conf))

        full_text = "\n".join([t[0] for t in lines])

        # 简单规则提取（PaddleOCR 不返回结构化发票数据，仅能做参考）
        import re

        invoice_number = ""
        invoice_date = ""
        total_amount = ""
        seller_name = ""

        # 发票号码：通常10位数字
        m = re.search(r"发票号码[:：]?\s*(\d{8,20})", full_text)
        if m:
            invoice_number = m.group(1)

        # 开票日期
        m = re.search(r"(\d{4}\s*年\s*\d{1,2}\s*月\s*\d{1,2}\s*日)", full_text)
        if m:
            invoice_date = _normalize_date(m.group(1))

        # 价税合计
        m = re.search(r"(?:价税合计|小写).*?[¥￥]\s*(\d+\.?\d*)", full_text)
        if m:
            total_amount = m.group(1)

        return {
            "invoice_number": invoice_number,
            "invoice_date": invoice_date,
            "seller_name": seller_name,
            "amount_excluding_tax": "",
            "tax_amount": "",
            "total_amount": total_amount,
            "description": "",
        }

    except Exception as exc:
        current_app.logger.warning("PaddleOCR exception: %s", exc)
        return None


@tool_bp.get("/reference")
def reference_data():
    """返回参考数据：产品线、办事处、部门、费用类别等。"""
    return jsonify(
        product_lines=PRODUCT_LINES,
        offices=OFFICES,
        departments=DEPARTMENTS,
        expense_categories=EXPENSE_CATEGORIES,
        customer_levels=CUSTOMER_LEVELS,
        entertainment_categories=ENTERTAINMENT_CATEGORIES,
    )


# ---------------------------------------------------------------------------
# 持久化存储：按费用期间保存/加载报销数据
# ---------------------------------------------------------------------------
def _rb_owner() -> tuple[str, str]:
    """获取当前用户的 owner 标识。
    优先用 X-RB-Anon-Id 头（前端 localStorage），回退到 Flask session 的 anon_id。
    """
    from auth.decorators import ensure_anon_id
    from flask_login import current_user

    if current_user.is_authenticated:
        return ("user", str(current_user.id))

    # Vercel serverless 会在冷启动时丢失会话 cookie——强制从客户端读取
    header_aid = (request.headers.get("X-RB-Anon-Id") or "").strip()
    if header_aid:
        return ("anon", header_aid)

    return ("anon", ensure_anon_id())


@tool_bp.post("/save")
@csrf.exempt
def save_state():
    """保存当前期间的完整报销状态。JSON: {period, data: {...}}"""
    payload = request.get_json(silent=True) or {}
    period = (payload.get("period") or "").strip()
    if not period:
        return jsonify(error="缺少费用期间"), 400

    otype, oid = _rb_owner()
    record = ReimbursementRecord.query.filter_by(
        owner_type=otype, owner_id=oid, period=period
    ).first()
    if not record:
        record = ReimbursementRecord(
            owner_type=otype, owner_id=oid, period=period,
            data_json="{}",
        )
        db.session.add(record)

    record.data_json = json.dumps(payload.get("data", {}), ensure_ascii=False)
    db.session.commit()
    return jsonify(success=True, period=period)


@tool_bp.get("/load/<period>")
def load_state(period: str):
    """加载指定期间的报销数据。"""
    period = (period or "").strip()
    if not period:
        return jsonify(success=False, note="期间名为空"), 400
    otype, oid = _rb_owner()
    record = ReimbursementRecord.query.filter_by(
        owner_type=otype, owner_id=oid, period=period
    ).first()
    if not record:
        return jsonify(success=False, note="该期间暂无数据")

    try:
        data = json.loads(record.data_json)
    except Exception:
        data = {}
    return jsonify(success=True, data=data, updated_at=record.updated_at.isoformat() if record.updated_at else None)


@tool_bp.get("/periods")
def list_periods():
    """列出当前用户所有期间及基本信息。"""
    otype, oid = _rb_owner()
    records = ReimbursementRecord.query.filter_by(
        owner_type=otype, owner_id=oid
    ).order_by(ReimbursementRecord.period.desc()).all()

    periods = []
    for r in records:
        try:
            d = json.loads(r.data_json)
            header = d.get("header", {})
            invoices = d.get("invoices", [])
            total = sum(_safe_float(inv.get("data", {}).get("total_amount")) for inv in invoices)
            periods.append({
                "period": r.period,
                "employee": header.get("employee_name", ""),
                "department": header.get("department", ""),
                "invoice_count": len(invoices),
                "total_amount": round(total, 2),
                "updated_at": r.updated_at.isoformat() if r.updated_at else None,
            })
        except Exception:
            periods.append({"period": r.period, "invoice_count": 0, "total_amount": 0})
    return jsonify(success=True, periods=periods)


@tool_bp.post("/delete-period")
@csrf.exempt
def delete_period():
    """删除指定期间的全部数据。JSON: {period}"""
    payload = request.get_json(silent=True) or {}
    period = (payload.get("period") or "").strip()
    if not period:
        return jsonify(success=False, error="期间名为空"), 400
    otype, oid = _rb_owner()
    record = ReimbursementRecord.query.filter_by(
        owner_type=otype, owner_id=oid, period=period
    ).first()
    if not record:
        # 调试：查找 DB 中是否真的有此 period 名（任何用户）
        any_match = ReimbursementRecord.query.filter_by(period=period).first()
        current_app.logger.warning(
            "delete_period 找不到记录: period=%r owner_type=%s owner_id=%s "
            "（DB 同名存在: %s）",
            period, otype, oid[:8] + '***', bool(any_match),
        )
        return jsonify(
            success=False,
            error="该期间不存在",
            debug={"period": period, "owner_type": otype, "owner_id_prefix": oid[:8]},
        ), 404
    db.session.delete(record)
    db.session.commit()
    return jsonify(success=True)


@tool_bp.post("/rename-period")
@csrf.exempt
def rename_period():
    """重命名期间。JSON: {old_period, new_period}"""
    payload = request.get_json(silent=True) or {}
    old = (payload.get("old_period") or "").strip()
    new = (payload.get("new_period") or "").strip()
    if not old or not new:
        return jsonify(success=False, error="参数不完整"), 400
    otype, oid = _rb_owner()
    record = ReimbursementRecord.query.filter_by(
        owner_type=otype, owner_id=oid, period=old
    ).first()
    if not record:
        return jsonify(success=False, error="原期间不存在"), 404
    # 检查新名称是否已存在
    conflict = ReimbursementRecord.query.filter_by(
        owner_type=otype, owner_id=oid, period=new
    ).first()
    if conflict:
        return jsonify(success=False, error="新期间名称已存在"), 409
    record.period = new
    db.session.commit()
    return jsonify(success=True, old_period=old, new_period=new)


@tool_bp.get("/debug")
def debug_info():
    """调试端点：返回当前 owner 和所有可见期间名。"""
    otype, oid = _rb_owner()
    periods = ReimbursementRecord.query.filter_by(
        owner_type=otype, owner_id=oid
    ).all()
    return jsonify(
        success=True,
        owner_type=otype,
        owner_id_prefix=oid[:8] + "***",
        my_periods=[p.period for p in periods],
        my_count=len(periods),
        total_periods=ReimbursementRecord.query.count(),
    )


@tool_bp.post("/cover-data")
@csrf.exempt
def cover_data():
    """
    接收前端传来的完整报销数据，返回封面所需的汇总计算结果。

    前端发送：
    {
        "header": {"employee_name":"","department":"","date":"","reason":"","period":""},
        "invoices": [{...发票字段..., "product_line":"","expense_type":"","office":"",
                       "customer_level":"","entertainment":{...}, "vehicle":{...}, "travel":{...} }]
    }
    后端返回汇总 + 中文大写金额。
    """
    data = request.get_json(silent=True) or {}
    header = data.get("header", {})
    invoices = data.get("invoices", [])

    # 按产品线+代码分组汇总
    groups: dict[str, dict] = {}
    for inv in invoices:
        pl_name = inv.get("product_line", "") or "未分类"
        pl_code = inv.get("product_line_code", "") or "-"
        office = inv.get("office", "") or ""
        key = f"{pl_code}|{pl_name}"

        total = _safe_float(inv.get("total_amount", 0))
        exp_type = inv.get("expense_type", "")

        if key not in groups:
            groups[key] = {
                "product_line": pl_name,
                "code": pl_code,
                "office": office,
                "totals": {c["key"]: 0.0 for c in EXPENSE_CATEGORIES},
                "remarks": [],
            }
        for cat in EXPENSE_CATEGORIES:
            if exp_type == cat["key"]:
                groups[key]["totals"][cat["key"]] += total
        rem = inv.get("remarks", "") or ""
        if rem and rem not in groups[key]["remarks"]:
            groups[key]["remarks"].append(rem)

    # 排序
    sorted_groups = sorted(groups.values(), key=lambda g: g["code"])

    # 合计
    grand_totals = {c["key"]: 0.0 for c in EXPENSE_CATEGORIES}
    for g in sorted_groups:
        for k in grand_totals:
            grand_totals[k] += g["totals"][k]

    total_all = sum(grand_totals.values())

    # 按费用类别汇总（封面左下角）
    expense_summary = [
        {
            "key": cat["key"],
            "label": cat["label"],
            "amount": round(grand_totals[cat["key"]], 2),
        }
        for cat in EXPENSE_CATEGORIES
    ]

    # 按客户等级汇总（费用分类表）— 仅计入有匹配费用类别的发票
    valid_exp_keys = {c["key"] for c in EXPENSE_CATEGORIES}
    level_groups: dict[str, dict] = {}
    for inv in invoices:
        level = inv.get("customer_level", "") or "未分类"
        total = _safe_float(inv.get("total_amount", 0))
        exp_type = inv.get("expense_type", "")
        if exp_type not in valid_exp_keys:
            continue  # 跳过无费用类别的发票，与封面 total_all 保持一致
        if level not in level_groups:
            level_groups[level] = {
                "level": level,
                "entertainment": 0.0,
                "travel": 0.0,
                "other": 0.0,
                "total": 0.0,
            }

        if exp_type in ("entertainment",):
            level_groups[level]["entertainment"] += total
        elif exp_type in ("travel_transport", "travel_hotel"):
            level_groups[level]["travel"] += total
        else:
            level_groups[level]["other"] += total
        level_groups[level]["total"] += total

    # 排序 level_groups 使结果顺序稳定
    level_groups = sorted(level_groups.values(), key=lambda g: g["level"])

    # 应酬费明细——优先取顶层数组（前端 genCover 发送），回退到发票内嵌数据
    entertainment_raw = data.get("entertainment") or []
    entertainment_details = []
    if entertainment_raw:
        for ent in entertainment_raw:
            amt = _safe_float(ent.get("amount"))
            if amt > 0:
                entertainment_details.append({
                    "date": ent.get("date", ""),
                    "category": ent.get("category", ""),
                    "place": ent.get("place", ""),
                    "customer": ent.get("customer", ""),
                    "participants": ent.get("participants", ""),
                    "amount": amt,
                    "purpose": ent.get("purpose", ""),
                })
    else:
        for inv in invoices:
            ent = inv.get("entertainment", {})
            if ent and ent.get("amount", ""):
                entertainment_details.append({
                    "date": ent.get("date", ""),
                    "category": ent.get("category", ""),
                    "place": ent.get("place", ""),
                    "customer": ent.get("customer", ""),
                    "participants": ent.get("participants", ""),
                    "amount": _safe_float(ent.get("amount")),
                    "purpose": ent.get("purpose", ""),
                })

    # 派车单明细——优先取顶层数组
    vehicles_raw = data.get("vehicles") or []
    vehicle_details = []
    if vehicles_raw:
        for veh in vehicles_raw:
            km = _safe_float(veh.get("km_total")); toll = _safe_float(veh.get("toll_fee")); parking = _safe_float(veh.get("parking_fee"))
            if km > 0 or toll > 0 or parking > 0:
                vehicle_details.append({
                    "date": veh.get("date", ""),
                    "from_location": veh.get("from_location", ""),
                    "to_location": veh.get("to_location", ""),
                    "contact": veh.get("contact", ""),
                    "km_start": veh.get("km_start", ""),
                    "km_end": veh.get("km_end", ""),
                    "km_total": km,
                    "toll_fee": toll,
                    "parking_fee": parking,
                    "remarks": veh.get("remarks", ""),
                })
    else:
        for inv in invoices:
            veh = inv.get("vehicle", {})
            km = _safe_float(veh.get("km_total")); toll = _safe_float(veh.get("toll_fee")); parking = _safe_float(veh.get("parking_fee"))
            if km > 0 or toll > 0 or parking > 0:
                vehicle_details.append({
                    "date": veh.get("date", ""),
                    "from_location": veh.get("from_location", ""),
                    "to_location": veh.get("to_location", ""),
                    "contact": veh.get("contact", ""),
                    "km_start": veh.get("km_start", ""),
                    "km_end": veh.get("km_end", ""),
                    "km_total": km,
                    "toll_fee": toll,
                    "parking_fee": parking,
                    "remarks": veh.get("remarks", ""),
                })

    # 出差明细——优先取顶层数组
    travels_raw = data.get("travels") or []
    travel_details = []
    if travels_raw:
        for tr in travels_raw:
            amt = _safe_float(tr.get("amount"))
            if amt > 0:
                travel_details.append({
                    "date": tr.get("date", ""),
                    "location": tr.get("location", ""),
                    "customer": tr.get("customer", ""),
                    "expense_type": tr.get("expense_type", ""),
                    "amount": amt,
                    "purpose": tr.get("purpose", ""),
                })
    else:
        for inv in invoices:
            tr = inv.get("travel", {})
            amt = _safe_float(tr.get("amount"))
            if tr and amt > 0:
                travel_details.append({
                    "date": tr.get("date", ""),
                    "location": tr.get("location", ""),
                    "customer": tr.get("customer", ""),
                    "expense_type": tr.get("expense_type", ""),
                    "amount": amt,
                    "purpose": tr.get("purpose", ""),
                })

    return jsonify(
        success=True,
        header=header,
        groups=sorted_groups,
        grand_totals={
            cat["key"]: round(grand_totals[cat["key"]], 2)
            for cat in EXPENSE_CATEGORIES
        },
        expense_summary=expense_summary,
        total_all=round(total_all, 2),
        total_cn=_number_to_chinese(round(total_all, 2)),
        invoice_count=len(invoices),
        level_groups=list(level_groups.values()),
        entertainment_details=entertainment_details,
        vehicle_details=vehicle_details,
        travel_details=travel_details,
        expense_categories=EXPENSE_CATEGORIES,
        customer_levels=CUSTOMER_LEVELS,
    )


# ---------------------------------------------------------------------------
# 模板下载
# ---------------------------------------------------------------------------
TEMPLATES = {
    "entertainment": "应酬费_出差明细_派车单_模板.xls",
    "cover": "报销封面及费用分类表_模板.xls",
}

@tool_bp.get("/templates")
def list_templates():
    """列出可下载的空白模板。"""
    return jsonify(
        templates=[
            {
                "id": "entertainment",
                "name": "应酬费、出差明细、派车单模板",
                "download": "/tools/reimbursement/download-template/entertainment",
            },
            {
                "id": "cover",
                "name": "报销封面及费用分类表模板",
                "download": "/tools/reimbursement/download-template/cover",
            },
        ]
    )


@tool_bp.get("/download-template/<template_id>")
def download_template(template_id):
    """下载空白模板文件。"""
    from flask import send_file

    filename = TEMPLATES.get(template_id)
    if not filename:
        return jsonify(error="模板不存在"), 404

    filepath = Path(current_app.config["UPLOAD_DIR"]) / "reimbursement_templates" / filename
    if not filepath.exists():
        return jsonify(error="模板文件不存在"), 404

    return send_file(
        str(filepath),
        mimetype="application/vnd.ms-excel",
        as_attachment=True,
        download_name=filename,
    )


# ---------------------------------------------------------------------------
# 导出 Excel（保持原格式）
# ---------------------------------------------------------------------------
def _excel_styles():
    """共享样式对象，避免重复创建。"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    thin = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))
    return {
        "thin": thin,
        "title": Font(name="微软雅黑", size=14, bold=True),
        "subtitle": Font(name="微软雅黑", size=12, bold=True),
        "header": Font(name="微软雅黑", size=10, bold=True),
        "normal": Font(name="微软雅黑", size=10),
        "bold": Font(name="微软雅黑", size=10, bold=True),
        "header_fill": PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"),
        "center": Alignment(horizontal="center", vertical="center", wrap_text=True),
        "left": Alignment(horizontal="left", vertical="center", wrap_text=True),
        "right": Alignment(horizontal="right", vertical="center"),
    }


def _build_cover_file(cover_data):
    """
    生成文件1：报销封面及费用分类表.xlsx
    严格按原始 Excel 行列布局（对照 xlrd 0-based → openpyxl 1-based）。
    """
    import io, openpyxl
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    s = _excel_styles()

    wb = openpyxl.Workbook()
    header = cover_data.get("header", {})
    groups = cover_data.get("groups", [])
    expense_cats = EXPENSE_CATEGORIES
    total_all = cover_data.get("total_all", 0)
    total_cn = cover_data.get("total_cn", "")
    EMP = header.get("employee_name", "")
    DEPT = header.get("department", "")
    DATE = header.get("date", "")
    PERIOD = header.get("period", "")

    # ========== Sheet 1: 封面 (费用报销单) ==========
    ws = wb.active
    ws.title = "封面"

    # Row 1 (原始 row 0): 标题 "费用报销单"
    ws.merge_cells("A1:O1")
    ws["A1"].value = "费用报销单"; ws["A1"].font = Font(name="微软雅黑", size=16, bold=True); ws["A1"].alignment = s["center"]
    ws.row_dimensions[1].height = 40

    # Rows 2-3: 空行
    ws.row_dimensions[2].height = 8
    ws.row_dimensions[3].height = 8

    # Row 4 (原始 row 3): 信息行 — B=员工姓名, G=部门, L=报销日期
    ws.merge_cells("B4:C4"); ws["B4"].value = f"员工姓名：{EMP}"; ws["B4"].font = s["normal"]
    ws.merge_cells("G4:H4"); ws["G4"].value = f"部门：{DEPT}"; ws["G4"].font = s["normal"]
    ws.merge_cells("L4:M4"); ws["L4"].value = f"报销日期：{DATE}"; ws["L4"].font = s["normal"]

    # Row 5: 空行
    ws.row_dimensions[5].height = 8

    # Row 6 (原始 row 5): 表头 — cols A-O
    col_headers = ["序号", "产品线", "代码", "办事处", "费用期间"]
    for cat in expense_cats:
        col_headers.append(cat["label"])
    col_headers.append("备注栏")

    thin = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))
    hfill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    hfont = Font(name="微软雅黑", size=10, bold=True)
    nfont = Font(name="微软雅黑", size=10)
    bfont = Font(name="微软雅黑", size=10, bold=True)

    for ci, h in enumerate(col_headers, 1):
        cell = ws.cell(row=6, column=ci, value=h)
        cell.font = hfont; cell.fill = hfill; cell.alignment = s["center"]; cell.border = thin
    ws.row_dimensions[6].height = 24

    # Rows 7-18 (原始 rows 6-17): 数据行
    row = 6
    for gi, g in enumerate(groups):
        row += 1
        vals = [gi + 1, g["product_line"], g["code"], g.get("office", ""), PERIOD]
        for cat in expense_cats:
            v = g["totals"].get(cat["key"], 0)
            vals.append(v if v != 0 else "")
        vals.append("；".join(g.get("remarks", [])))
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=ci, value=v)
            cell.font = nfont; cell.border = thin
            cell.alignment = s["right"] if isinstance(v, (int, float)) and v != "" else s["center"]

    # Row 19 (原始 row 18): 合计行 — "合计"在 E 列(第5列)，后面跟各费用类别合计
    row += 1
    # 所有列都画边框
    for ci in range(1, len(col_headers) + 1):
        ws.cell(row=row, column=ci).border = thin
    # "合计" 在 E 列 (费用期间列)
    ws.cell(row=row, column=5).value = "合计"; ws.cell(row=row, column=5).font = bfont; ws.cell(row=row, column=5).alignment = s["center"]
    # 各费用类别合计（从 F 列开始）
    for ci, cat in enumerate(expense_cats, 6):
        v = cover_data["grand_totals"].get(cat["key"], 0)
        cell = ws.cell(row=row, column=ci, value=v if v != 0 else "")
        cell.font = bfont; cell.alignment = s["right"]; cell.border = thin

    # Row 20: 空行
    row += 1

    # Row 21 (原始 row 20): 费用总额（小写）— F=费用总额, G=（小写）, H=值
    row += 1
    ws.cell(row=row, column=6).value = "费用总额："; ws.cell(row=row, column=6).font = nfont
    ws.cell(row=row, column=7).value = "（小写）"; ws.cell(row=row, column=7).font = nfont
    ws.cell(row=row, column=8).value = round(total_all, 2); ws.cell(row=row, column=8).font = nfont

    # Row 22 (原始 row 21): 费用总额（大写）— F=费用总额, G=（大写）, H=中文大写
    row += 1
    ws.cell(row=row, column=6).value = "费用总额："; ws.cell(row=row, column=6).font = nfont
    ws.cell(row=row, column=7).value = "（大写）"; ws.cell(row=row, column=7).font = nfont
    ws.cell(row=row, column=8).value = total_cn; ws.cell(row=row, column=8).font = nfont

    # Row 23 (原始 row 22): 空行
    row += 1

    # Row 24 (原始 row 23): 签字行 — B=申请人, F=财务审核, I=部门主管, M=总经理
    row += 1
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
    ws.cell(row=row, column=2).value = f"申请人:{EMP}"; ws.cell(row=row, column=2).font = nfont
    ws.cell(row=row, column=6).value = "财务审核:"; ws.cell(row=row, column=6).font = nfont
    ws.cell(row=row, column=9).value = "部门主管:"; ws.cell(row=row, column=9).font = nfont
    ws.cell(row=row, column=13).value = "总经理："; ws.cell(row=row, column=13).font = nfont

    # Row 25 (原始 row 24): 日期行 — B=日期, F=日期, I=日期, M=日期
    row += 1
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
    ws.cell(row=row, column=2).value = f"日期：{DATE}"; ws.cell(row=row, column=2).font = nfont
    ws.cell(row=row, column=6).value = "日期："; ws.cell(row=row, column=6).font = nfont
    ws.cell(row=row, column=9).value = "日期："; ws.cell(row=row, column=9).font = nfont
    ws.cell(row=row, column=13).value = "日期："; ws.cell(row=row, column=13).font = nfont

    # 列宽
    col_widths = [5, 16, 5, 7, 9] + [9] * len(expense_cats) + [14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ========== Sheet 2: 费用分类表 ==========
    ws2 = wb.create_sheet("费用分类表")

    # Row 1: 标题
    ws2.merge_cells("A1:H1")
    ws2["A1"].value = "按客户等级费用分类明细表"; ws2["A1"].font = Font(name="微软雅黑", size=14, bold=True); ws2["A1"].alignment = s["center"]
    ws2.row_dimensions[1].height = 36

    # Row 2 (原始 row 1): 表头 — A=序号, B=客户等级, C=招待费, D=差旅费, E=空, F=其他费用, G=费用合计, H=备注
    cls_h = ["序号", "客户等级", "招待费", "差旅费", "", "其他费用", "费用合计", "备注"]
    for ci, h in enumerate(cls_h, 1):
        cell = ws2.cell(row=2, column=ci, value=h)
        cell.font = hfont; cell.fill = hfill; cell.alignment = s["center"]; cell.border = thin

    # Rows 3-6 (原始 rows 2-5): 数据行（4 levels，固定顺序）
    level_groups = cover_data.get("level_groups", [])
    # 按 0-1, level 1, level 2, level 3 固定顺序排列
    lvl_order = {"0-1": 0, "level 1": 1, "level 2": 2, "level 3": 3}
    level_groups.sort(key=lambda lg: lvl_order.get(lg.get("level", ""), 99))

    row = 2
    for li, lg in enumerate(level_groups):
        row += 1
        vs = [li + 1, lg.get("level", ""),
              lg.get("entertainment", 0) or "", lg.get("travel", 0) or "", "",
              lg.get("other", 0) or "", lg.get("total", 0) or "", ""]
        for ci, v in enumerate(vs, 1):
            cell = ws2.cell(row=row, column=ci, value=v)
            cell.font = nfont; cell.border = thin
            cell.alignment = s["right"] if isinstance(v, (int, float)) else s["center"]

    # Row 7: 空行
    row += 1

    # Row 8 (原始 row 7): 费用总额（小写）— C=费用总额, D=（小写）, E=值
    row += 1
    ws2.cell(row=row, column=3).value = "费用总额："; ws2.cell(row=row, column=3).font = nfont
    ws2.cell(row=row, column=4).value = "（小写）"; ws2.cell(row=row, column=4).font = nfont
    ws2.cell(row=row, column=5).value = round(total_all, 2); ws2.cell(row=row, column=5).font = nfont

    # Row 9 (原始 row 8): 费用总额（大写）— C=费用总额, D=（大写）, E=中文大写
    row += 1
    ws2.cell(row=row, column=3).value = "费用总额："; ws2.cell(row=row, column=3).font = nfont
    ws2.cell(row=row, column=4).value = "（大写）"; ws2.cell(row=row, column=4).font = nfont
    ws2.cell(row=row, column=5).value = total_cn; ws2.cell(row=row, column=5).font = nfont

    # Row 10: 空行
    row += 1

    # Row 11 (原始 row 10): 签字 — B=申请人, F=部门主管
    row += 1
    ws2.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
    ws2.cell(row=row, column=2).value = f"申请人:{EMP}"; ws2.cell(row=row, column=2).font = nfont
    ws2.cell(row=row, column=6).value = "部门主管:"; ws2.cell(row=row, column=6).font = nfont

    # Row 12 (原始 row 11): 日期 — B=日期, F=日期
    row += 1
    ws2.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
    ws2.cell(row=row, column=2).value = f"日期：{DATE}"; ws2.cell(row=row, column=2).font = nfont
    ws2.cell(row=row, column=6).value = "日期："; ws2.cell(row=row, column=6).font = nfont

    # 列宽
    for i, w in enumerate([6, 12, 12, 12, 4, 12, 12, 12], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return buf


def _build_detail_file(cover_data, entertainment, vehicles, travels):
    """
    生成文件2：应酬费_出差明细_派车单.xlsx
    三张表：应酬费明细表 + 派车单 + 出差明细表
    严格按原始 Excel 行列布局（对照 xlrd 0-based → openpyxl 1-based）。
    """
    import io, openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    s = _excel_styles()

    thin = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))
    hfill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    hfont = Font(name="微软雅黑", size=10, bold=True)
    nfont = Font(name="微软雅黑", size=10)
    bfont = Font(name="微软雅黑", size=10, bold=True)
    tfont = Font(name="微软雅黑", size=14, bold=True)

    header = cover_data.get("header", {})
    EMP = header.get("employee_name", "")
    PERIOD = header.get("period", "")

    wb = openpyxl.Workbook()

    # ========== Sheet 1: 应酬费明细表 ==========
    ws = wb.active
    ws.title = "应酬费明细"

    # Row 1: 标题
    ws.merge_cells("A1:H1")
    ws["A1"].value = "应酬费报销明细"; ws["A1"].font = tfont; ws["A1"].alignment = s["center"]

    # Row 2: 员工姓名（不合并，直接放 A2）
    ws.cell(row=2, column=1, value=f"员工姓名：{EMP}").font = nfont

    # Row 3: 表头
    eh = ["时间", "费用类别", "用餐地点、名称", "客户名称", "主要参与人全称", "金额", "事由（请注明是 主要推广哪条产品线）", "主管审批（超金额）"]
    for ci, h in enumerate(eh, 1):
        cell = ws.cell(row=3, column=ci, value=h)
        cell.font = hfont; cell.fill = hfill; cell.alignment = s["center"]; cell.border = thin

    # Rows 4+: 数据
    row = 3
    ent_total = 0
    for e in entertainment:
        row += 1
        amt = _safe_float(e.get("amount"))
        vals = [e.get("date", ""), e.get("category", ""), e.get("place", ""),
                e.get("customer", ""), e.get("participants", ""), amt or "",
                e.get("purpose", ""), ""]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=ci, value=v)
            cell.font = nfont; cell.border = thin
            cell.alignment = s["right"] if ci == 6 else s["left"]
        ent_total += amt

    # 合计行：只在 F 列（col 6）放总金额，无"合计"标签
    row += 1
    for ci in range(1, 9):
        ws.cell(row=row, column=ci).border = thin
    ws.cell(row=row, column=6).value = round(ent_total, 2) if ent_total else ""; ws.cell(row=row, column=6).font = bfont; ws.cell(row=row, column=6).alignment = s["right"]

    for i, w in enumerate([12, 10, 18, 14, 18, 10, 18, 12], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ========== Sheet 2: 派车单 ==========
    ws2 = wb.create_sheet("派车单")

    # Row 1: 公司名
    ws2.merge_cells("A1:J1")
    ws2["A1"].value = "威欣电子有限公司"; ws2["A1"].font = tfont; ws2["A1"].alignment = s["center"]

    # Row 2: 员工姓名 + 期间（J 列右对齐）
    ws2.cell(row=2, column=1, value=f"员工姓名：{EMP}").font = nfont
    period_text = f" {PERIOD}派车单" if PERIOD else "派车单"
    ws2.cell(row=2, column=10, value=period_text).font = nfont
    ws2.cell(row=2, column=10).alignment = Alignment(horizontal="right")

    # Row 3: 表头 — "行车路程"合并 D:E，客户联系人在 F
    ws2.merge_cells(start_row=3, start_column=4, end_row=3, end_column=5)
    ws2.cell(row=3, column=4).value = "行车路程"; ws2.cell(row=3, column=4).font = hfont; ws2.cell(row=3, column=4).fill = hfill; ws2.cell(row=3, column=4).alignment = s["center"]; ws2.cell(row=3, column=4).border = thin

    # 列顺序: A=日期 B=出发地 C=目的地 D:E=行车路程 F=客户联系人 G=公里数 H=过桥费 I=停车费 J=备注
    vh1 = ["日期", "出发地", "目的地", None, "客户联系人", "公里数", "过桥费", "停车费", "备注"]
    for ci, h in enumerate(vh1, 1):
        if h is None: continue  # 跳过 D:E 合并区域
        cell = ws2.cell(row=3, column=ci, value=h)
        cell.font = hfont; cell.fill = hfill; cell.alignment = s["center"]; cell.border = thin

    # Row 4: 子表头 — D4=起, E4=止
    ws2.cell(row=4, column=4).value = "起"; ws2.cell(row=4, column=4).font = hfont; ws2.cell(row=4, column=4).fill = hfill; ws2.cell(row=4, column=4).alignment = s["center"]; ws2.cell(row=4, column=4).border = thin
    ws2.cell(row=4, column=5).value = "止"; ws2.cell(row=4, column=5).font = hfont; ws2.cell(row=4, column=5).fill = hfill; ws2.cell(row=4, column=5).alignment = s["center"]; ws2.cell(row=4, column=5).border = thin

    # Rows 5+: 数据 — D=km_start, E=km_end, F=contact
    row = 4
    km_t, toll_t, park_t = 0, 0, 0
    for v in vehicles:
        row += 1
        km = _safe_float(v.get("km_total")); toll = _safe_float(v.get("toll_fee")); park = _safe_float(v.get("parking_fee"))
        vals = [v.get("date", ""), v.get("from_location", ""), v.get("to_location", ""),
                v.get("km_start", ""), v.get("km_end", ""), v.get("contact", ""),
                km or "", toll or "", park or "", v.get("remarks", "")]
        for ci, val in enumerate(vals, 1):
            cell = ws2.cell(row=row, column=ci, value=val)
            cell.font = nfont; cell.border = thin
            cell.alignment = s["right"] if ci >= 7 else s["left"]
        km_t += km; toll_t += toll; park_t += park

    # 合计行 — A=合计, G=公里数合计, H=过桥费合计, I=停车费合计
    row += 1
    for ci in range(1, 11):
        ws2.cell(row=row, column=ci).border = thin
    ws2.cell(row=row, column=1).value = "合计"; ws2.cell(row=row, column=1).font = bfont; ws2.cell(row=row, column=1).alignment = s["center"]
    ws2.cell(row=row, column=7).value = round(km_t, 1) if km_t else ""; ws2.cell(row=row, column=7).font = bfont; ws2.cell(row=row, column=7).alignment = s["right"]
    ws2.cell(row=row, column=8).value = round(toll_t, 2) if toll_t else ""; ws2.cell(row=row, column=8).font = bfont; ws2.cell(row=row, column=8).alignment = s["right"]
    ws2.cell(row=row, column=9).value = round(park_t, 2) if park_t else ""; ws2.cell(row=row, column=9).font = bfont; ws2.cell(row=row, column=9).alignment = s["right"]

    for i, w in enumerate([12, 8, 18, 12, 8, 8, 8, 8, 8, 14], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # ========== Sheet 3: 出差明细表 ==========
    ws3 = wb.create_sheet("出差明细")

    # Row 1: 标题
    ws3.merge_cells("A1:F1")
    ws3["A1"].value = "出差明细表"; ws3["A1"].font = tfont; ws3["A1"].alignment = s["center"]

    # Row 2: 员工姓名
    ws3.cell(row=2, column=1, value=f"员工姓名：{EMP}").font = nfont

    # Row 3: 表头
    th = ["日期", "出差地", "客户名称", "费用类型", "金额", "事由（请注明是 主要推广哪条产品线）"]
    for ci, h in enumerate(th, 1):
        cell = ws3.cell(row=3, column=ci, value=h)
        cell.font = hfont; cell.fill = hfill; cell.alignment = s["center"]; cell.border = thin

    # Rows 4+: 数据
    row = 3
    tr_total = 0
    for t in travels:
        row += 1
        amt = float(t.get("amount", 0) or 0)
        vals = [t.get("date", ""), t.get("location", ""), t.get("customer", ""),
                t.get("expense_type", ""), amt or "", t.get("purpose", "")]
        for ci, v in enumerate(vals, 1):
            cell = ws3.cell(row=row, column=ci, value=v)
            cell.font = nfont; cell.border = thin
            cell.alignment = s["right"] if ci == 5 else s["left"]
        tr_total += amt

    # 合计行：只在有数据时添加
    if tr_total > 0:
        row += 1
        for ci in range(1, 7):
            ws3.cell(row=row, column=ci).border = thin
        ws3.cell(row=row, column=5).value = round(tr_total, 2); ws3.cell(row=row, column=5).font = bfont; ws3.cell(row=row, column=5).alignment = s["right"]

    for i, w in enumerate([12, 14, 14, 10, 10, 20], 1):
        ws3.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return buf


@tool_bp.post("/export-cover")
@csrf.exempt
def export_cover():
    """导出文件1：报销封面及费用分类表.xlsx（2张表：封面 + 费用分类表）"""
    from flask import send_file
    data = request.get_json(silent=True) or {}
    cover_data = data.get("cover_data", {})
    if not cover_data:
        return jsonify(error="无封面数据"), 400
    try:
        buf = _build_cover_file(cover_data)
        emp = _safe_filename(cover_data.get("header", {}).get("employee_name", "报销"))
        period = _safe_filename(cover_data.get("header", {}).get("period", ""))
        filename = f"报销封面及费用分类表_{emp}_{period}.xlsx"
        return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         as_attachment=True, download_name=filename)
    except Exception as e:
        current_app.logger.exception("export-cover failed")
        return jsonify(error=f"导出失败：{str(e)}"), 500


@tool_bp.post("/export-details")
@csrf.exempt
def export_details():
    """导出文件2：应酬费_出差明细_派车单.xlsx（3张表：应酬费明细 + 派车单 + 出差明细）"""
    from flask import send_file
    data = request.get_json(silent=True) or {}
    cover_data = data.get("cover_data", {})
    entertainment = data.get("entertainment") or []
    vehicles = data.get("vehicles") or []
    travels = data.get("travels") or []
    if not cover_data:
        return jsonify(error="无数据"), 400
    try:
        buf = _build_detail_file(cover_data, entertainment, vehicles, travels)
        emp = _safe_filename(cover_data.get("header", {}).get("employee_name", "报销"))
        period = _safe_filename(cover_data.get("header", {}).get("period", ""))
        filename = f"应酬费_出差明细_派车单_{emp}_{period}.xlsx"
        return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         as_attachment=True, download_name=filename)
    except Exception as e:
        current_app.logger.exception("export-details failed")
        return jsonify(error=f"导出失败：{str(e)}"), 500
