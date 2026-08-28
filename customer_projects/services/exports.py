"""Bounded, audited Excel exports for the customer-project ledger."""
from __future__ import annotations

import hashlib
from dataclasses import dataclass
from datetime import datetime, timezone
from io import BytesIO
from typing import Any, Mapping

from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Font, PatternFill
from sqlalchemy import func, select

from customer_projects.models import Customer, CustomerProject, ProjectExportPolicy, ProjectMaterial, ProjectStatusCatalog
from customer_projects.permissions import can_edit_prices
from customer_projects.services.projects import add_audit
from extensions import db
from models import User
from shared.models import OrganizationMembership


DEFAULT_EXPORT_ROLES = frozenset({"organization_admin", "business_manager"})


class ControlledExportError(ValueError):
    pass


@dataclass(frozen=True)
class ProjectExportArtifact:
    content: bytes
    filename: str
    sha256: str
    project_count: int
    material_count: int
    row_count: int
    includes_prices: bool


def ensure_default_export_policy(organization_id: str) -> ProjectExportPolicy:
    policy = db.session.scalar(
        select(ProjectExportPolicy).where(ProjectExportPolicy.organization_id == organization_id)
    )
    if policy is None:
        policy = ProjectExportPolicy(organization_id=organization_id)
        policy.set_allowed_roles(DEFAULT_EXPORT_ROLES)
        db.session.add(policy)
        db.session.flush()
    return policy


def export_allowed(
    membership: OrganizationMembership, policy: ProjectExportPolicy | None
) -> bool:
    allowed_roles = policy.allowed_roles if policy is not None else DEFAULT_EXPORT_ROLES
    return bool(membership.roles.intersection(allowed_roles))


def _safe_excel_value(value: Any) -> Any:
    if not isinstance(value, str):
        return value
    cleaned = ILLEGAL_CHARACTERS_RE.sub("", value)[:32767]
    if cleaned.lstrip().startswith(("=", "+", "-", "@")):
        return "'" + cleaned
    return cleaned


def _audit_filters(filters: Mapping[str, str]) -> dict[str, str]:
    return {
        "q": str(filters.get("q") or "")[:100],
        "stage": str(filters.get("stage") or "")[:32],
    }


def _aware(value: datetime) -> datetime:
    return value.replace(tzinfo=timezone.utc) if value.tzinfo is None else value.astimezone(timezone.utc)


def build_project_export(
    statement: Any,
    membership: OrganizationMembership,
    policy: ProjectExportPolicy,
    filters: Mapping[str, str],
) -> ProjectExportArtifact:
    if not export_allowed(membership, policy):
        raise ControlledExportError("当前角色没有项目导出权限。")

    projects = list(
        db.session.scalars(
            statement.order_by(CustomerProject.updated_at.desc(), CustomerProject.id)
            .limit(policy.max_projects + 1)
        )
    )
    if len(projects) > policy.max_projects:
        add_audit(
            membership.organization_id,
            "customer_project_export",
            membership.organization_id,
            "blocked",
            membership.user_id,
            {
                "reason": "project_limit_exceeded",
                "max_projects": policy.max_projects,
                "filters": _audit_filters(filters),
                "policy_version": policy.version,
            },
        )
        raise ControlledExportError(
            f"当前筛选结果超过 {policy.max_projects} 个项目，请缩小筛选范围后重试。"
        )

    project_ids = [project.id for project in projects]
    material_counts = {
        project_id: count
        for project_id, count in db.session.execute(
            select(ProjectMaterial.project_id, func.count(ProjectMaterial.id))
            .where(
                ProjectMaterial.organization_id == membership.organization_id,
                ProjectMaterial.project_id.in_(project_ids),
                ProjectMaterial.deleted_at.is_(None),
            )
            .group_by(ProjectMaterial.project_id)
        )
    } if project_ids else {}
    material_count = sum(int(count) for count in material_counts.values())
    row_count = sum(max(1, int(material_counts.get(project_id, 0))) for project_id in project_ids)
    if row_count > policy.max_rows:
        add_audit(
            membership.organization_id,
            "customer_project_export",
            membership.organization_id,
            "blocked",
            membership.user_id,
            {
                "reason": "row_limit_exceeded",
                "max_rows": policy.max_rows,
                "project_count": len(projects),
                "filters": _audit_filters(filters),
                "policy_version": policy.version,
            },
        )
        raise ControlledExportError(
            f"当前筛选结果将生成超过 {policy.max_rows} 行，请缩小筛选范围后重试。"
        )

    materials = list(
        db.session.scalars(
            select(ProjectMaterial)
            .where(
                ProjectMaterial.organization_id == membership.organization_id,
                ProjectMaterial.project_id.in_(project_ids),
                ProjectMaterial.deleted_at.is_(None),
            )
            .order_by(ProjectMaterial.project_id, ProjectMaterial.is_primary.desc(), ProjectMaterial.id)
        )
    ) if project_ids else []
    materials_by_project: dict[str, list[ProjectMaterial]] = {project_id: [] for project_id in project_ids}
    for material in materials:
        materials_by_project[material.project_id].append(material)

    customer_ids = {project.customer_id for project in projects}
    customers = {
        customer.id: customer
        for customer in db.session.scalars(select(Customer).where(Customer.id.in_(customer_ids)))
    } if customer_ids else {}
    owner_ids = {project.primary_sales_user_id for project in projects}
    owners = {
        user.id: user.display_name
        for user in db.session.scalars(select(User).where(User.id.in_(owner_ids)))
    } if owner_ids else {}
    stages = {
        row.code: row.display_name
        for row in db.session.scalars(
            select(ProjectStatusCatalog).where(
                ProjectStatusCatalog.organization_id == membership.organization_id
            )
        )
    }

    includes_prices = bool(policy.include_prices and can_edit_prices(membership))
    headers = [
        "项目编号", "客户", "客户评级", "项目名称", "产品名称", "项目年用量", "阶段",
        "主业务", "下一步", "下次跟进", "推广品牌", "推广型号", "应用位置", "单机数量",
    ]
    if includes_prices:
        headers.extend(["录入单价", "录入币别", "美元单价", "含税人民币单价", "USD/CNY汇率", "价格更新时间"])

    workbook = Workbook(write_only=True)
    sheet = workbook.create_sheet("客户项目台账")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{'T' if includes_prices else 'N'}{row_count + 1}"
    header_cells = []
    for header in headers:
        cell = WriteOnlyCell(sheet, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="0D6EFD")
        header_cells.append(cell)
    sheet.append(header_cells)

    for project in projects:
        customer = customers.get(project.customer_id)
        for material in materials_by_project.get(project.id) or [None]:
            row = [
                project.project_code,
                customer.name if customer else "",
                customer.grade if customer else "",
                project.name,
                project.product_name or "",
                float(project.annual_usage) if project.annual_usage is not None else None,
                stages.get(project.stage_code, project.stage_code),
                owners.get(project.primary_sales_user_id, ""),
                project.next_action,
                _aware(project.next_follow_up_at).strftime("%Y-%m-%d %H:%M"),
                material.promoted_brand if material else "",
                (material.promoted_mpn or "待确认") if material else "",
                (material.application_position or "") if material else "",
                float(material.machine_quantity) if material and material.machine_quantity is not None else None,
            ]
            if includes_prices:
                row.extend([
                    float(material.target_price) if material and material.target_price is not None else None,
                    (material.currency or "") if material else "",
                    float(material.unit_price_usd) if material and material.unit_price_usd is not None else None,
                    float(material.unit_price_cny_tax_included) if material and material.unit_price_cny_tax_included is not None else None,
                    float(material.fx_rate_usd_cny) if material and material.fx_rate_usd_cny is not None else None,
                    _aware(material.price_updated_at).strftime("%Y-%m-%d %H:%M") if material and material.price_updated_at else "",
                ])
            sheet.append([_safe_excel_value(value) for value in row])

    output = BytesIO()
    workbook.save(output)
    content = output.getvalue()
    digest = hashlib.sha256(content).hexdigest()
    add_audit(
        membership.organization_id,
        "customer_project_export",
        membership.organization_id,
        "exported",
        membership.user_id,
        {
            "project_count": len(projects),
            "material_count": material_count,
            "row_count": row_count,
            "includes_prices": includes_prices,
            "filters": _audit_filters(filters),
            "file_sha256": digest,
            "policy_version": policy.version,
        },
    )
    return ProjectExportArtifact(
        content=content,
        filename=f"客户项目台账-{datetime.now().strftime('%Y%m%d')}.xlsx",
        sha256=digest,
        project_count=len(projects),
        material_count=material_count,
        row_count=row_count,
        includes_prices=includes_prices,
    )
