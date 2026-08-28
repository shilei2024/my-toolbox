"""Bounded Excel preview, commit and conditional revert for project migration."""
from __future__ import annotations

import hashlib
import json
import unicodedata
import zipfile
from datetime import date, datetime, time, timezone
from decimal import Decimal, InvalidOperation
from io import BytesIO

from openpyxl import load_workbook
from sqlalchemy import func, select
from werkzeug.utils import secure_filename

from extensions import db
from models import User
from shared.models import Organization, OrganizationMembership

from customer_projects.models import (
    Customer,
    CustomerProject,
    ProjectImportBatch,
    ProjectImportRow,
    ProjectStatusCatalog,
)
from customer_projects.services.projects import (
    DomainError,
    add_audit,
    create_customer,
    create_project,
    normalize_name,
    parse_datetime,
    soft_delete_project,
)


MAX_IMPORT_BYTES = 5 * 1024 * 1024
MAX_IMPORT_ROWS = 1000
MAX_XLSX_MEMBERS = 500
MAX_XLSX_UNCOMPRESSED_BYTES = 25 * 1024 * 1024
REQUIRED_FIELDS = (
    "customer_name",
    "project_name",
    "product_name",
    "annual_usage",
    "stage_code",
    "owner_email",
    "next_action",
    "next_follow_up_at",
)
HEADER_ALIASES = {
    "customer_name": {"客户名称", "客户", "customer_name"},
    "project_name": {"项目名称", "项目", "project_name"},
    "product_name": {"产品名称", "产品", "product_name"},
    "annual_usage": {"项目年用量", "年用量", "annual_usage"},
    "stage_code": {"阶段", "当前阶段", "stage_code"},
    "owner_email": {"主负责人邮箱", "主业务邮箱", "负责人邮箱", "owner_email"},
    "next_action": {"下一步", "下一步动作", "next_action"},
    "next_follow_up_at": {"下次跟进时间", "下次跟进", "next_follow_up_at"},
    "assessment_grade": {"项目等级", "评估等级", "assessment_grade"},
    "probability_band": {"成功概率", "成功概率区间", "probability_band"},
}


class ProjectImportError(ValueError):
    pass


def _header_key(value: object) -> str:
    return unicodedata.normalize("NFKC", str(value or "")).strip().casefold().replace(" ", "")


def _validate_xlsx_archive(content: bytes) -> None:
    if len(content) > MAX_IMPORT_BYTES:
        raise ProjectImportError("Excel 文件不能超过 5MB。")
    try:
        with zipfile.ZipFile(BytesIO(content)) as archive:
            members = archive.infolist()
            if len(members) > MAX_XLSX_MEMBERS:
                raise ProjectImportError("Excel 内部文件数量异常。")
            total = 0
            for member in members:
                if member.flag_bits & 0x1:
                    raise ProjectImportError("不支持加密 Excel。")
                total += member.file_size
                if total > MAX_XLSX_UNCOMPRESSED_BYTES:
                    raise ProjectImportError("Excel 解压后内容超过 25MB。")
                if member.compress_size and member.file_size / member.compress_size > 200:
                    raise ProjectImportError("Excel 压缩比异常。")
    except zipfile.BadZipFile as exc:
        raise ProjectImportError("文件不是有效的 XLSX。") from exc


def _json_value(value: object) -> object:
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return str(value)
    return value


def _parse_follow_up(value: object, timezone_name: str) -> datetime:
    if isinstance(value, datetime):
        return parse_datetime(value, timezone_name)
    if isinstance(value, date):
        return parse_datetime(datetime.combine(value, time(9, 0)), timezone_name)
    return parse_datetime(str(value or ""), timezone_name)


def preview_project_import(
    content: bytes,
    filename: str,
    membership: OrganizationMembership,
) -> ProjectImportBatch:
    if not filename.lower().endswith(".xlsx"):
        raise ProjectImportError("仅支持 .xlsx 模板文件。")
    _validate_xlsx_archive(content)
    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:  # openpyxl exposes several parser exception types
        raise ProjectImportError("Excel 无法读取或结构已损坏。") from exc
    try:
        sheet = workbook.active
        iterator = sheet.iter_rows(values_only=True)
        headers = next(iterator, None)
        if not headers:
            raise ProjectImportError("Excel 缺少表头。")
        alias_lookup = {
            _header_key(alias): field
            for field, aliases in HEADER_ALIASES.items()
            for alias in aliases
        }
        mapping: dict[str, int] = {}
        source_headers: dict[str, str] = {}
        for index, header in enumerate(headers):
            field = alias_lookup.get(_header_key(header))
            if field and field not in mapping:
                mapping[field] = index
                source_headers[field] = str(header or "")
        missing = [field for field in REQUIRED_FIELDS if field not in mapping]
        if missing:
            raise ProjectImportError("模板缺少必填列：" + "、".join(missing))

        raw_rows: list[tuple[int, tuple[object, ...]]] = []
        for row_number, values in enumerate(iterator, start=2):
            if not any(value not in (None, "") for value in values):
                continue
            raw_rows.append((row_number, values))
            if len(raw_rows) > MAX_IMPORT_ROWS:
                raise ProjectImportError("单个批次最多导入 1000 行。")
    finally:
        workbook.close()

    organization = db.session.get(Organization, membership.organization_id)
    timezone_name = organization.timezone if organization else "Asia/Shanghai"
    statuses = list(
        db.session.scalars(
            select(ProjectStatusCatalog).where(
                ProjectStatusCatalog.organization_id == membership.organization_id,
                ProjectStatusCatalog.is_active.is_(True),
            )
        )
    )
    stage_lookup = {
        _header_key(value): item.code
        for item in statuses
        for value in (item.code, item.display_name)
    }
    owner_cache: dict[str, int | None] = {}
    customer_cache: dict[str, list[Customer]] = {}
    seen_projects: set[tuple[str, str]] = set()
    batch = ProjectImportBatch(
        organization_id=membership.organization_id,
        original_filename=secure_filename(filename)[:255] or "projects.xlsx",
        file_sha256=hashlib.sha256(content).hexdigest(),
        status="preview",
        mapping_json=json.dumps(source_headers, ensure_ascii=False, sort_keys=True),
        created_by_user_id=membership.user_id,
    )
    db.session.add(batch)
    db.session.flush()

    valid_count = error_count = 0
    for row_number, values in raw_rows:
        payload = {
            field: _json_value(values[index] if index < len(values) else None)
            for field, index in mapping.items()
        }
        errors: list[str] = []
        for field in REQUIRED_FIELDS:
            if payload.get(field) in (None, ""):
                errors.append(f"{source_headers[field]}必填")
        customer_name = str(payload.get("customer_name") or "").strip()
        project_name = str(payload.get("project_name") or "").strip()
        normalized_customer = normalize_name(customer_name)
        normalized_project = normalize_name(project_name)
        payload["customer_name"] = customer_name
        payload["project_name"] = project_name
        payload["product_name"] = str(payload.get("product_name") or "").strip()
        payload["next_action"] = str(payload.get("next_action") or "").strip()

        try:
            annual_usage = Decimal(str(payload.get("annual_usage") or ""))
            if annual_usage <= 0 or annual_usage != annual_usage.to_integral_value():
                raise InvalidOperation
            payload["annual_usage"] = str(annual_usage.to_integral_value())
        except (InvalidOperation, ValueError):
            errors.append("项目年用量必须是大于 0 的整数")
        stage = stage_lookup.get(_header_key(payload.get("stage_code")))
        if stage not in {"evaluation", "initiated", "sampling", "pilot_batch", "trial_production", "design_win"}:
            errors.append("阶段必须是有效的进行中阶段")
        else:
            payload["stage_code"] = stage
        try:
            payload["next_follow_up_at"] = _parse_follow_up(
                payload.get("next_follow_up_at"), timezone_name
            ).isoformat()
        except DomainError:
            errors.append("下次跟进时间格式无效")

        email = str(payload.get("owner_email") or "").strip().casefold()
        payload["owner_email"] = email
        if email not in owner_cache:
            owner_membership = db.session.scalar(
                select(OrganizationMembership)
                .join(User, OrganizationMembership.user_id == User.id)
                .where(
                    User.email == email,
                    User.is_active_user.is_(True),
                    OrganizationMembership.organization_id == membership.organization_id,
                    OrganizationMembership.status == "active",
                )
            )
            owner_cache[email] = (
                owner_membership.user_id
                if owner_membership
                and owner_membership.roles.intersection(
                    {"organization_admin", "business_manager", "sales"}
                )
                else None
            )
        owner_id = owner_cache[email]
        if owner_id is None:
            errors.append("主负责人邮箱不是当前组织的有效成员")
        else:
            payload["primary_sales_user_id"] = owner_id

        if normalized_customer not in customer_cache:
            customer_cache[normalized_customer] = list(
                db.session.scalars(
                    select(Customer).where(
                        Customer.organization_id == membership.organization_id,
                        Customer.normalized_name == normalized_customer,
                        Customer.deleted_at.is_(None),
                    )
                )
            )
        customers = customer_cache[normalized_customer]
        if len(customers) > 1:
            errors.append("组织内存在多个同名客户，需先清理")
        elif customers:
            payload["existing_customer_id"] = customers[0].id
            existing_project = db.session.scalar(
                select(CustomerProject.id).where(
                    CustomerProject.organization_id == membership.organization_id,
                    CustomerProject.customer_id == customers[0].id,
                    CustomerProject.normalized_name == normalized_project,
                    CustomerProject.deleted_at.is_(None),
                )
            )
            if existing_project:
                errors.append("该客户下已存在同名项目")
        duplicate_key = (normalized_customer, normalized_project)
        if duplicate_key in seen_projects:
            errors.append("批次内存在重复客户与项目名称")
        seen_projects.add(duplicate_key)

        grade = str(payload.get("assessment_grade") or "").strip().upper()
        if grade and grade not in {"A", "B", "C", "D"}:
            errors.append("项目等级仅支持 A/B/C/D")
        payload["assessment_grade"] = grade or None
        probability = str(payload.get("probability_band") or "").strip().rstrip("%")
        if probability:
            try:
                probability_value = int(float(probability))
            except ValueError:
                probability_value = 0
            if probability_value not in {10, 30, 50, 70, 90}:
                errors.append("成功概率仅支持 10/30/50/70/90")
            else:
                payload["probability_band"] = probability_value
        else:
            payload["probability_band"] = None

        status = "invalid" if errors else "valid"
        valid_count += status == "valid"
        error_count += status == "invalid"
        db.session.add(
            ProjectImportRow(
                organization_id=membership.organization_id,
                batch_id=batch.id,
                row_number=row_number,
                status=status,
                payload_json=json.dumps(payload, ensure_ascii=False, sort_keys=True),
                errors_json=json.dumps(errors, ensure_ascii=False),
            )
        )
    batch.total_rows = len(raw_rows)
    batch.valid_rows = valid_count
    batch.error_rows = error_count
    add_audit(
        membership.organization_id,
        "project_import_batch",
        batch.id,
        "previewed",
        membership.user_id,
        {"rows": batch.total_rows, "valid": valid_count, "errors": error_count, "file_sha256": batch.file_sha256},
    )
    return batch


def commit_project_import(
    batch: ProjectImportBatch, membership: OrganizationMembership
) -> ProjectImportBatch:
    if batch.organization_id != membership.organization_id:
        raise ProjectImportError("导入批次不存在。")
    if batch.status == "committed":
        return batch
    if batch.status != "preview":
        raise ProjectImportError("当前批次不能确认导入。")
    rows = list(
        db.session.scalars(
            select(ProjectImportRow)
            .where(ProjectImportRow.batch_id == batch.id, ProjectImportRow.status == "valid")
            .order_by(ProjectImportRow.row_number)
        )
    )
    if not rows:
        raise ProjectImportError("批次没有可导入的有效行。")
    customer_cache: dict[str, tuple[Customer, bool]] = {}
    for row in rows:
        payload = json.loads(row.payload_json)
        customer_key = normalize_name(payload["customer_name"])
        if customer_key not in customer_cache:
            customer = (
                db.session.get(Customer, payload.get("existing_customer_id"))
                if payload.get("existing_customer_id")
                else None
            )
            was_created = False
            if customer is None or customer.deleted_at is not None:
                customer = create_customer(
                    {
                        "name": payload["customer_name"],
                        "primary_owner_user_id": payload["primary_sales_user_id"],
                    },
                    membership,
                )
                was_created = True
            customer_cache[customer_key] = (customer, was_created)
        customer, was_created = customer_cache[customer_key]
        project_payload = dict(payload)
        project_payload["customer_id"] = customer.id
        project_payload["name"] = payload["project_name"]
        project = create_project(
            project_payload,
            membership,
            f"import:{batch.id}:{row.row_number}",
        )
        row.customer_id = customer.id
        row.project_id = project.id
        row.customer_was_created = was_created
        row.project_version_at_create = project.version
        row.status = "created"
    batch.status = "committed"
    batch.committed_at = datetime.now(timezone.utc)
    add_audit(
        membership.organization_id,
        "project_import_batch",
        batch.id,
        "committed",
        membership.user_id,
        {"created_projects": len(rows), "invalid_rows_skipped": batch.error_rows},
    )
    return batch


def revert_project_import(
    batch: ProjectImportBatch, membership: OrganizationMembership
) -> dict[str, int]:
    if batch.organization_id != membership.organization_id:
        raise ProjectImportError("导入批次不存在。")
    if batch.status not in {"committed", "partially_reverted"}:
        raise ProjectImportError("当前批次不能撤销。")
    rows = list(
        db.session.scalars(
            select(ProjectImportRow).where(
                ProjectImportRow.batch_id == batch.id,
                ProjectImportRow.status.in_(("created", "not_revertible")),
            )
        )
    )
    reverted = blocked = 0
    created_customer_ids: set[str] = set()
    for row in rows:
        project = db.session.get(CustomerProject, row.project_id) if row.project_id else None
        if (
            project is None
            or project.deleted_at is not None
            or project.version != row.project_version_at_create
        ):
            row.status = "not_revertible"
            blocked += 1
            continue
        soft_delete_project(
            project,
            f"撤销导入批次 {batch.id}",
            membership,
            project.version,
        )
        row.status = "reverted"
        reverted += 1
        if row.customer_was_created and row.customer_id:
            created_customer_ids.add(row.customer_id)
    db.session.flush()
    for customer_id in created_customer_ids:
        customer = db.session.get(Customer, customer_id)
        active_projects = db.session.scalar(
            select(func.count()).select_from(CustomerProject).where(
                CustomerProject.customer_id == customer_id,
                CustomerProject.deleted_at.is_(None),
            )
        ) or 0
        if customer and customer.deleted_at is None and customer.version == 1 and active_projects == 0:
            customer.deleted_at = datetime.now(timezone.utc)
            customer.deleted_by_user_id = membership.user_id
            customer.delete_reason = f"撤销导入批次 {batch.id}"
            customer.version += 1
            customer.updated_by_user_id = membership.user_id
            add_audit(
                membership.organization_id,
                "customer",
                customer.id,
                "deleted",
                membership.user_id,
                {"reason": "import_batch_reverted", "batch_id": batch.id},
            )
    batch.status = "partially_reverted" if blocked else "reverted"
    batch.reverted_at = datetime.now(timezone.utc)
    add_audit(
        membership.organization_id,
        "project_import_batch",
        batch.id,
        "reverted",
        membership.user_id,
        {"reverted_projects": reverted, "blocked_projects": blocked},
    )
    return {"reverted": reverted, "blocked": blocked}
